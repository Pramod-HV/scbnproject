#!/usr/bin/env python3
"""
Script to create project documentation in Excel format with descriptions
"""

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

def create_project_documentation():
    """Create Excel file with task list and descriptions"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Tasks"
    
    # Define styles
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 70
    
    # Add headers
    headers = ['', 'Task #', 'Task Name', 'Description']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Project tasks with descriptions
    tasks = [
        ("", "5.1", "Create Formal Excel Structure for Checklist", 
         "Designed and created structured Excel checklist (Generic_checklistMain.xlsm) with multiple sheets for configuration management"),
        
        ("", "5.2", "Develop Map Features Extraction Utility", 
         "Built utility to extract map details from MXL files and store in Excel: Map Name, Direction (Inbound/Outbound), Input/Output Format Classification, Document Type and Format Detection, SAP Classification, Base Name"),
        
        ("", "5.3", "Implement Map Renaming Logic", 
         "Developed logic to rename MXL files based on ZIP file names and update the Description tag within MXL files to match new names"),
        
        ("", "5.4", "Check Document Type and Direction for Updates", 
         "Implemented validation to match Direction and Document Type from mapping_results.xlsx with process_data_updates sheet before processing"),
        
        ("", "5.5", "Develop ExplicitRule Insertion for MXL Processing", 
         "Created core functionality to insert ExplicitRule tags in MXL files with update processdata statements for BusinessReference and ExtraInfo fields"),
        
        ("", "5.6", "Support NON-EDI Formats (XML/Positional)", 
         "Extended processing to support NON-EDI formats including XML hierarchical search and Positional format field mapping"),
        
        ("", "5.7", "Integrate Character Encoding Modification", 
         "Added feature to modify character encoding in MXL files based on Generic_rule sheet configuration (UTF-8, ISO-8859-1, etc.)"),
        
        ("", "5.8", "Integrate X Syntax Token Modification", 
         "Implemented X syntax token modification feature to update token values in MXL files based on requirements"),
        
        ("", "5.9", "Develop Comment/Uncomment Preservation", 
         "Built two-phase system to extract comments before processing and restore them after, preventing loss of documentation"),
        
        ("", "5.10", "Integrate Freeformat Processing", 
         "Added support for freeformat processing in MXL files based on Generic_rule sheet configuration"),
        
        ("", "5.11", "Integrate PreSession Rule Updates", 
         "Implemented PreSessionRule section updates in MXL files during Stage 2 processing"),
        
        ("", "5.12", "Create Two-Stage Process Orchestrator", 
         "Developed orchestrator script (process_all_mxl_files.py) with Stage 1 (extract comments) and Stage 2 (apply changes, restore comments)"),
        
        ("", "5.13", "Develop Error Reporting System", 
         "Created comprehensive error reporting system that generates Excel reports with deduplication for inactive fields and mapping errors"),
        
        ("", "5.14", "Implement Inactive Field Handling", 
         "Added fallback logic to find next active field when target field is inactive, with error reporting when no active fields exist"),
        
        ("", "5.15", "Add MapName Column Feature", 
         "Implemented MapName column support in process_data_updates sheet for better traceability of which map each rule applies to"),
        
        ("", "5.16", "Develop SAP IDOC Field Processing", 
         "Built feature to process SAP IDOC fields in MXL files based on Inbound_maps sheet configuration (rows 5-13)"),
        
        ("", "5.17", "Implement Positional Delimiter Configuration", 
         "Added feature to configure positional delimiters (CR, LF, CRLF) in PosSyntax section from Inbound_maps sheet (rows 15-16)"),
        
        ("", "5.18", "Develop XML Output Configuration", 
         "Implemented XML output settings configuration: prolog, doctype, formatting, character entity references from Inbound_maps sheet (rows 18-22)"),
        
        ("", "5.19", "Develop CSV Output Configuration", 
         "Added CSV output configuration feature: field delimiter, quote character, column names from Inbound_maps sheet (rows 24-28)"),
        
        ("", "5.20", "Create Project Documentation", 
         "Developed comprehensive project documentation including usage guides, quick reference, and this task documentation"),
    ]
    
    # Add tasks to worksheet
    row = 2
    for task in tasks:
        for col, value in enumerate(task, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            if col == 4:  # Description column
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        row += 1
    
    # Save workbook
    wb.save("Project_Tasks_Documentation.xlsx")
    print("Project documentation created successfully: Project_Tasks_Documentation.xlsx")

if __name__ == "__main__":
    create_project_documentation()

# Made with Bob
