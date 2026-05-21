#!/usr/bin/env python3
"""
Modify Syntax Token Script
Modifies the X syntax token in MXL files based on Generic_rule sheet.
"""

import xml.etree.ElementTree as ET
import openpyxl
import os
import re
from typing import Optional, Tuple


def load_generic_rules(excel_file: str) -> dict:
    """
    Load Generic_rule sheet from Excel file.
    
    Args:
        excel_file: Path to the Excel file
        
    Returns:
        Dictionary mapping (direction, document_type) to X syntax token value
    """
    rules = {}
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        if 'Generic_rule' not in wb.sheetnames:
            print(f"Warning: 'Generic_rule' sheet not found in {excel_file}")
            return rules
            
        ws = wb['Generic_rule']
        
        # Read headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            if cell_value:
                headers[cell_value.strip()] = col
        
        # Check required columns
        required_cols = ['Direction', 'Document Type', 'X syntax token']
        for col_name in required_cols:
            if col_name not in headers:
                print(f"Warning: Required column '{col_name}' not found in Generic_rule sheet")
                return rules
        
        # Read data rows
        for row in range(2, ws.max_row + 1):
            direction = ws.cell(row, headers['Direction']).value
            doc_type = ws.cell(row, headers['Document Type']).value
            x_token = ws.cell(row, headers['X syntax token']).value
            
            if direction and doc_type and x_token:
                key = (str(direction).strip().upper(), str(doc_type).strip().upper())
                rules[key] = str(x_token).strip()
        
        print(f"Loaded {len(rules)} rules from Generic_rule sheet")
        
    except Exception as e:
        print(f"Error loading Generic_rule sheet: {e}")
    
    return rules


def modify_x_syntax_token(mxl_file: str, x_token_value: str) -> bool:
    """
    Modify the X syntax token in an MXL file.
    
    Args:
        mxl_file: Path to the MXL file
        x_token_value: New value for X syntax token
        
    Returns:
        True if modification was successful, False otherwise
    """
    try:
        # Read the file content
        with open(mxl_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Find and replace X syntax token
        # Pattern: <SyntaxToken>X</SyntaxToken> where X can be any value
        pattern = r'(<SyntaxToken>)[^<]+(</SyntaxToken>)'
        
        # Check if pattern exists
        if not re.search(pattern, content):
            print(f"Warning: No SyntaxToken found in {mxl_file}")
            return False
        
        # Replace with new value
        new_content = re.sub(pattern, rf'\1{x_token_value}\2', content)
        
        # Write back to file
        with open(mxl_file, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        print(f"✓ Modified X syntax token to '{x_token_value}' in {os.path.basename(mxl_file)}")
        return True
        
    except Exception as e:
        print(f"Error modifying {mxl_file}: {e}")
        return False


def get_map_info_from_filename(filename: str) -> Optional[Tuple[str, str]]:
    """
    Extract direction and document type from MXL filename.
    
    Args:
        filename: Name of the MXL file
        
    Returns:
        Tuple of (direction, document_type) or None if not found
    """
    # Pattern: direction_format_doctype
    # Example: "Inbound_EDI(X12, 850, 004010)_Positional.mxl"
    
    filename_upper = filename.upper()
    
    # Determine direction
    direction = None
    if 'INBOUND' in filename_upper or '_I_' in filename_upper:
        direction = 'INBOUND'
    elif 'OUTBOUND' in filename_upper or '_O_' in filename_upper:
        direction = 'OUTBOUND'
    
    # Extract document type
    doc_type = None
    
    # Try to find EDI document type (e.g., 850, 856, 204)
    edi_match = re.search(r'[_\s](\d{3})[_\s]', filename)
    if edi_match:
        doc_type = edi_match.group(1)
    
    # Try to find EDIFACT document type (e.g., ORDERS, DESADV)
    edifact_match = re.search(r'[_\s](ORDERS|DESADV|INVOIC|ORDRSP)[_\s]', filename_upper)
    if edifact_match:
        doc_type = edifact_match.group(1)
    
    if direction and doc_type:
        return (direction, doc_type)
    
    return None


def process_mxl_file(mxl_file: str, generic_rules: dict) -> bool:
    """
    Process a single MXL file and modify its X syntax token if applicable.
    
    Args:
        mxl_file: Path to the MXL file
        generic_rules: Dictionary of generic rules
        
    Returns:
        True if processed successfully, False otherwise
    """
    filename = os.path.basename(mxl_file)
    
    # Get map info from filename
    map_info = get_map_info_from_filename(filename)
    
    if not map_info:
        print(f"Could not extract direction/document type from {filename}")
        return False
    
    direction, doc_type = map_info
    key = (direction, doc_type)
    
    # Check if rule exists
    if key not in generic_rules:
        print(f"No rule found for {direction} {doc_type} in {filename}")
        return False
    
    x_token_value = generic_rules[key]
    
    # Modify the file
    return modify_x_syntax_token(mxl_file, x_token_value)


def process_all_mxl_files(mxl_directory: str, excel_file: str):
    """
    Process all MXL files in a directory.
    
    Args:
        mxl_directory: Directory containing MXL files
        excel_file: Path to the Excel file with Generic_rule sheet
    """
    print("=" * 60)
    print("Modifying X Syntax Tokens")
    print("=" * 60)
    
    # Load generic rules
    generic_rules = load_generic_rules(excel_file)
    
    if not generic_rules:
        print("No generic rules loaded. Exiting.")
        return
    
    # Process each MXL file
    mxl_files = [f for f in os.listdir(mxl_directory) if f.endswith('.mxl')]
    
    if not mxl_files:
        print(f"No MXL files found in {mxl_directory}")
        return
    
    print(f"\nFound {len(mxl_files)} MXL files to process\n")
    
    success_count = 0
    for mxl_file in mxl_files:
        full_path = os.path.join(mxl_directory, mxl_file)
        if process_mxl_file(full_path, generic_rules):
            success_count += 1
    
    print("\n" + "=" * 60)
    print(f"Processing complete: {success_count}/{len(mxl_files)} files modified")
    print("=" * 60)


if __name__ == "__main__":
    # Default paths
    MXL_DIRECTORY = "mxl_files"
    EXCEL_FILE = "Generic_checklist.xlsm"
    
    # Check if files exist
    if not os.path.exists(MXL_DIRECTORY):
        print(f"Error: Directory '{MXL_DIRECTORY}' not found")
        exit(1)
    
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file '{EXCEL_FILE}' not found")
        exit(1)
    
    # Process all files
    process_all_mxl_files(MXL_DIRECTORY, EXCEL_FILE)

# Made with Bob
