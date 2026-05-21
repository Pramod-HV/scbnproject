#!/usr/bin/env python3
"""
Inbound Maps Features Processor
Handles positional delimiter configuration from Inbound_maps sheet in Generic_checklistMain.xlsm
"""

import sys
import os
import re
import logging
from datetime import datetime
import openpyxl

def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def read_positional_config(checklist_file):
    """
    Read positional configuration from Inbound_maps sheet
    Returns: dict with 'is_positional', 'delimiter1', 'delimiter2'
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Inbound_maps']
        
        # Row 15, Column B: Is this a Positional? (Yes/No)
        is_positional_cell = ws.cell(row=15, column=2).value
        is_positional = str(is_positional_cell).strip().lower() if is_positional_cell else "no"
        
        # Row 15, Column C: Delimiter 1 (format: "Delimiter 1: CR")
        delimiter1_cell = ws.cell(row=15, column=3).value
        delimiter1 = ""
        if delimiter1_cell and str(delimiter1_cell).strip():
            cell_text = str(delimiter1_cell).strip()
            # Extract delimiter value after colon (e.g., "Delimiter 1: CR" -> "CR")
            if ':' in cell_text:
                delimiter1 = cell_text.split(':', 1)[1].strip()
            else:
                delimiter1 = cell_text
        
        # Row 16, Column C: Delimiter 2 (format: "Delimiter 2: LF")
        delimiter2_cell = ws.cell(row=16, column=3).value
        delimiter2 = ""
        if delimiter2_cell and str(delimiter2_cell).strip():
            cell_text = str(delimiter2_cell).strip()
            # Extract delimiter value after colon (e.g., "Delimiter 2: LF" -> "LF")
            if ':' in cell_text:
                delimiter2 = cell_text.split(':', 1)[1].strip()
            else:
                delimiter2 = cell_text
        
        wb.close()
        
        config = {
            'is_positional': is_positional == 'yes',
            'delimiter1': delimiter1,
            'delimiter2': delimiter2
        }
        
        logger.info(f"Positional config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading positional config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def read_syntax_record_config(checklist_file):
    """
    Read syntax record configuration from Inbound_maps sheet
    Returns: dict with 'syntax_type', 'delimiter_mode', and parsed offset values
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=True)
        ws = wb['Inbound_maps']
        
        # Row 2, Column B: Syntax Record dropdown (Yes(X12), Yes(EDIFACT), No)
        syntax_cell = ws.cell(row=2, column=2).value
        syntax_value = str(syntax_cell).strip().lower() if syntax_cell else "no"
        
        # Row 2, Column C: Delimiter mode (Remove Delimiter and add syntax record, Retain Delimiter and add syntax record)
        delimiter_cell = ws.cell(row=2, column=3).value
        delimiter_mode = str(delimiter_cell).strip().lower() if delimiter_cell else "retain"
        
        # Determine if we should remove or retain delimiters
        remove_delimiters = 'remove' in delimiter_mode
        
        # Parse the value
        if 'x12' in syntax_value:
            syntax_type = 'x12'
            # Read from D2 (column 4) for X12
            values_cell = ws.cell(row=2, column=4).value
        elif 'edifact' in syntax_value:
            syntax_type = 'edifact'
            # Read from E2 (column 5) for EDIFACT
            values_cell = ws.cell(row=2, column=5).value
        else:
            syntax_type = 'no'
            values_cell = None
        
        wb.close()
        
        # Parse the values from the cell
        offsets = None
        if values_cell and syntax_type != 'no':
            offsets = parse_syntax_record_values(values_cell, syntax_type)
        
        config = {
            'syntax_type': syntax_type,
            'remove_delimiters': remove_delimiters,
            'offsets': offsets
        }
        
        logger.info(f"Syntax Record config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading syntax record config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def parse_syntax_record_values(cell_value, syntax_type):
    """
    Parse syntax record values from Excel cell
    Returns dict with offset values
    """
    logger = logging.getLogger(__name__)
    
    try:
        lines = str(cell_value).strip().split('\n')
        values = {}
        tag_name = None
        
        for line in lines:
            if ':' in line:
                parts = line.split(':', 1)
                key = parts[0].strip()  # Keep original case
                value = parts[1].strip()
                
                # Special handling for 'tag' - first occurrence is the tag name (ISA/UNA)
                if key.lower() == 'tag' and tag_name is None:
                    tag_name = value
                    continue
                
                # Try to convert to int
                try:
                    value = int(value)
                except:
                    pass
                
                # Store with lowercase key for case-insensitive lookup
                values[key.lower()] = value
        
        # Use the tag name we found, or default
        if tag_name is None:
            tag_name = 'ISA' if syntax_type == 'x12' else 'UNA'
        
        # Map to XML tag names with position conversion
        # Helper function to get value with fallback keys
        def get_value(keys, default):
            for key in keys:
                if key in values and isinstance(values[key], int):
                    return values[key]
            return default
        
        # Helper function to convert position values
        # Rule: 0 -> 65535 (not used), 1 -> 0 (first position), 65535 -> 65535 (already not used), other -> value - 1
        def convert_position(value):
            if value == 0:
                return 65535
            elif value == 1:
                return 0
            elif value == 65535:
                return 65535
            else:
                return value - 1
        
        # Get position value and apply conversion
        tag_offset_val = int(values.get('position', 1))
        
        offsets = {
            'tag': tag_name,
            'contains_data': 'yes' if str(values.get('contains data', values.get('conatins data', 'yes'))).lower() in ['yes', 'checked'] else 'no',
            'tag_offset': convert_position(tag_offset_val),
            'length': int(values.get('length', 107 if syntax_type == 'x12' else 9)),
            'tag_delim_offset': convert_position(get_value(['tag position', 'tag'], 3)),
            'segment_delim_offset': convert_position(get_value(['segment position', 'segment'], 105)),
            'element_delim_offset': convert_position(get_value(['element position', 'element'], 3)),
            'repeat_delim_offset': convert_position(get_value(['repeating element position', 'repeating element'], 65535)),
            'subelement_delim_offset': convert_position(get_value(['subelement position', 'subelement'], 104)),
            'rel_char_offset': convert_position(get_value(['release character position', 'release character'], 65535)),
            'decimal_sep_offset': convert_position(get_value(['decimal seperator position', 'decimal seperator', 'decimal separator position', 'decimal separator'], 65535))
        }
        
        logger.info(f"Parsed offsets: {offsets}")
        return offsets
        
    except Exception as e:
        logger.error(f"Error parsing syntax record values: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def read_xml_output_config(checklist_file):
    """
    Read XML output configuration from Inbound_maps sheet (rows 18-22)
    Returns: dict with XML output settings or None if XML processing is disabled
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Inbound_maps']
        
        # Row 18, Column B: Is this a XML? (Yes/No)
        is_xml_cell = ws.cell(row=18, column=2).value
        is_xml = str(is_xml_cell).strip().lower() if is_xml_cell else "no"
        
        if is_xml != 'yes':
            logger.info("XML output processing is disabled (Row 18 = No)")
            wb.close()
            return None
        
        logger.info("XML output processing is enabled (Row 18 = Yes)")
        
        # Row 19, Column B: XML Prolog and Document type declaration
        prolog_cell = ws.cell(row=19, column=2).value
        prolog_option = str(prolog_cell).strip() if prolog_cell else ""
        
        # Map Excel dropdown values to MXL tag values
        output_header = "no_header"  # default
        if "no prolog" in prolog_option.lower():
            output_header = "no_header"
        elif "prolog and document type declaration specified" in prolog_option.lower():
            output_header = "prolog_and_doctype"
        elif "prolog specified" in prolog_option.lower():
            output_header = "prolog"
        
        # Row 19, Column C: publicID (only if prolog_and_doctype)
        public_id = ""
        if output_header == "prolog_and_doctype":
            public_id_cell = ws.cell(row=19, column=3).value
            public_id = str(public_id_cell).strip() if public_id_cell else ""
        
        # Row 20, Column C: systemID (only if prolog_and_doctype)
        system_id = ""
        if output_header == "prolog_and_doctype":
            system_id_cell = ws.cell(row=20, column=3).value
            system_id = str(system_id_cell).strip() if system_id_cell else ""
        
        # Row 21, Column B: Document Formatting
        format_cell = ws.cell(row=21, column=2).value
        format_option = str(format_cell).strip() if format_cell else ""
        
        # Map Excel dropdown values to MXL tag values
        # Check most specific patterns first
        output_format = "no_newlines"  # default
        
        format_lower = format_option.lower()
        # Check for "no indentation" FIRST before checking for "indent"
        if "no indentation" in format_lower:
            # "One element per line, no indentation"
            output_format = "newlines"
        elif "indent" in format_lower or "hierarchy" in format_lower:
            # "One element per line, indented to show element hierarchy"
            output_format = "newlines_and_indent"
        elif "no newlines" in format_lower:
            # "No newlines"
            output_format = "no_newlines"
        elif "newlines" in format_lower:
            # Fallback for just "newlines"
            output_format = "newlines"
        
        # Row 22, Column B: Use Character Entity References
        char_entity_cell = ws.cell(row=22, column=2).value
        char_entity = str(char_entity_cell).strip().lower() if char_entity_cell else "no"
        use_char_entity = "yes" if char_entity == "yes" else "no"
        
        wb.close()
        
        config = {
            'is_xml': True,
            'output_header': output_header,
            'public_id': public_id,
            'system_id': system_id,
            'output_format': output_format,
            'use_char_entity_ref': use_char_entity
        }
        
        logger.info(f"XML output config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading XML output config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
def read_csv_output_config(checklist_file):
    """
    Read CSV output configuration from Inbound_maps sheet (rows 24-28)
    Returns: dict with 'is_csv', 'field_delimiter', 'quote_character', 'include_column_names'
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Inbound_maps']
        
        # Row 24, Column B: Is this a CSV? (Yes/No)
        is_csv_cell = ws.cell(row=24, column=2).value
        is_csv = str(is_csv_cell).strip().lower() if is_csv_cell else "no"
        
        # If not CSV, return early
        if is_csv != 'yes':
            wb.close()
            return {
                'is_csv': False,
                'field_delimiter': None,
                'quote_character': None,
                'include_column_names': None
            }
        
        # Row 25, Column B: Delimiter Type (Default/Custom)
        delimiter_type_cell = ws.cell(row=25, column=2).value
        delimiter_type = str(delimiter_type_cell).strip().lower() if delimiter_type_cell else "default"
        
        # Determine field delimiter and quote character
        if delimiter_type == "default":
            field_delimiter = ","
            quote_character = '"'
        else:  # Custom
            # Row 26, Column B: Field Delimiter
            field_delimiter_cell = ws.cell(row=26, column=2).value
            field_delimiter = str(field_delimiter_cell).strip() if field_delimiter_cell else ","
            
            # Row 27, Column B: Quote Character
            quote_character_cell = ws.cell(row=27, column=2).value
            quote_character = str(quote_character_cell).strip() if quote_character_cell else '"'
        
        # Row 28, Column B: Include column names (Yes/No)
        include_names_cell = ws.cell(row=28, column=2).value
        include_names = str(include_names_cell).strip().lower() if include_names_cell else "no"
        include_column_names = "1" if include_names == "yes" else "0"
        
        wb.close()
        
        config = {
            'is_csv': True,
            'field_delimiter': field_delimiter,
            'quote_character': quote_character,
            'include_column_names': include_column_names
        }
        
        logger.info(f"CSV output config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading CSV output config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def read_sap_idoc_config(checklist_file):
    """
    Read SAP IDOC configuration from Inbound_maps sheet (rows 5-13)
    Returns: list of dicts with field configurations
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Inbound_maps']
        
        # Check if SAP IDOC processing is enabled (need to find the row)
        # Assuming it's in a specific row - need to locate "If client format is SAP IDOC"
        is_sap_idoc = False
        sap_idoc_row = None
        
        # Search for "If client format is SAP IDOC" in column A
        for row in range(1, 20):  # Search first 20 rows
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'SAP IDOC' in str(cell_value).upper():
                # Check column B for Yes/No
                answer = ws.cell(row=row, column=2).value
                if answer and str(answer).strip().lower() == 'yes':
                    is_sap_idoc = True
                    sap_idoc_row = row
                    logger.info(f"SAP IDOC processing enabled at row {row}")
                break
        
        if not is_sap_idoc:
            logger.info("SAP IDOC processing is disabled or not found")
            wb.close()
            return None
        
        # Read configurations from rows 5-13 (A5:E13)
        field_configs = []
        for row in range(5, 14):  # Rows 5 to 13
            col_a = ws.cell(row=row, column=1).value  # Field description
            col_b = ws.cell(row=row, column=2).value  # Dropdown (Text1-9 or None)
            col_c = ws.cell(row=row, column=3).value  # COLUMN_C value
            col_d = ws.cell(row=row, column=4).value  # COLUMN_D value
            col_e = ws.cell(row=row, column=5).value  # ISA/GS reference
            
            # Skip if column B is None or empty
            if not col_b or str(col_b).strip().lower() == 'none':
                continue
            
            # Extract field name from column A (e.g., "Add EDI_DC40 fields -- SNDPRN" -> "SNDPRN")
            field_name = None
            if col_a and '--' in str(col_a):
                field_name = str(col_a).split('--')[-1].strip()
            
            if field_name:
                config = {
                    'field_name': field_name,
                    'text_value': str(col_b).strip(),
                    'column_c': str(col_c).strip() if col_c else '',
                    'column_d': str(col_d).strip() if col_d else '',
                    'column_e': str(col_e).strip() if col_e else ''
                }
                field_configs.append(config)
                logger.info(f"Found SAP IDOC field config: {config}")
        
        wb.close()
        return field_configs
        
    except Exception as e:
        logger.error(f"Error reading SAP IDOC config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def find_edi_field_name(mxl_content, segment_tag, field_number):
    """
    Find the field name for a specific EDI segment and field number
    Example: ISA tag, 6th field -> returns field name like "IAO34"
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Find the segment tag (e.g., <Tag>ISA</Tag>)
        segment_pattern = rf'<Tag>{segment_tag}</Tag>.*?</Segment>'
        segment_match = re.search(segment_pattern, mxl_content, re.DOTALL)
        
        if not segment_match:
            logger.warning(f"Could not find segment tag: {segment_tag}")
            return None
        
        segment_content = segment_match.group(0)
        
        # Find all fields in this segment
        field_pattern = r'<Field>.*?</Field>'
        fields = re.findall(field_pattern, segment_content, re.DOTALL)
        
        if field_number > len(fields):
            logger.warning(f"Field number {field_number} exceeds available fields in {segment_tag}")
            return None
        
        # Get the specific field (1-based index)
        target_field = fields[field_number - 1]
        
        # Extract the Name tag value
        name_match = re.search(r'<Name>([^<]+)</Name>', target_field)
        if name_match:
            field_name = name_match.group(1)
            logger.info(f"Found field name '{field_name}' for {segment_tag}-{field_number:02d}")
            return field_name
        
        return None
        
    except Exception as e:
        logger.error(f"Error finding EDI field name: {e}")
        return None

def build_column_d_expression(mxl_content, column_e):
    """
    Build the COLUMN_D expression based on column E value
    Examples:
    - "ISA06" -> "ISA06"  (find ISA tag, 6th field name)
    - "ISA06-GS02" -> "IAO34 + \"_\" + G142"  (find both field names)
    """
    logger = logging.getLogger(__name__)
    
    if not column_e:
        return ""
    
    try:
        # Check if it contains a hyphen (multiple fields)
        if '-' in column_e:
            parts = column_e.split('-')
            field_names = []
            
            for part in parts:
                # Extract segment and field number (e.g., "ISA06" -> "ISA", 6)
                segment = ''.join([c for c in part if c.isalpha()])
                field_num_str = ''.join([c for c in part if c.isdigit()])
                
                if segment and field_num_str:
                    field_num = int(field_num_str)
                    field_name = find_edi_field_name(mxl_content, segment, field_num)
                    if field_name:
                        field_names.append(field_name)
            
            if field_names:
                # Add # prefix to each field name and join with + "_" +
                prefixed_fields = [f'#{name}' for name in field_names]
                expression = ' + "_" + '.join(prefixed_fields)
                logger.info(f"Built expression from '{column_e}': {expression}")
                return expression
        else:
            # Single field reference (e.g., "ISA06")
            segment = ''.join([c for c in column_e if c.isalpha()])
            field_num_str = ''.join([c for c in column_e if c.isdigit()])
            
            if segment and field_num_str:
                field_num = int(field_num_str)
                field_name = find_edi_field_name(mxl_content, segment, field_num)
                if field_name:
                    # Add # prefix to field name
                    prefixed_name = f'#{field_name}'
                    logger.info(f"Built expression from '{column_e}': {prefixed_name}")
                    return prefixed_name
        
        return ""
        
    except Exception as e:
        logger.error(f"Error building COLUMN_D expression: {e}")
        return ""

def process_sap_idoc_field(mxl_content, field_config):
    """
    Process a single SAP IDOC field configuration
    - Find the field by name (e.g., <Name>SNDPRN</Name>)
    - Empty <Link> and <ImplicitRule> tags
    - Comment out existing <ExplicitRule> content
    - Add new ExplicitRule with codelist query
    """
    logger = logging.getLogger(__name__)
    
    try:
        field_name = field_config['field_name']
        text_value = field_config['text_value']
        column_c = field_config['column_c']
        column_d = field_config['column_d']
        column_e = field_config['column_e']
        
        logger.info(f"Processing SAP IDOC field: {field_name}")
        
        # Find ALL fields that match - must match <Name> tag exactly at the start of the field
        # Pattern: <Field> followed by any content, then <Name>FIELDNAME</Name> or <Name>FIELDNAME:XX</Name>
        # Use non-greedy match and ensure we're matching the Name tag right after Field opening
        field_pattern = rf'<Field>\s*<ID>\d+</ID>\s*<Name>{re.escape(field_name)}(?::\d+)?</Name>.*?</Field>'
        field_matches = list(re.finditer(field_pattern, mxl_content, re.DOTALL))
        
        if not field_matches:
            logger.warning(f"Could not find any fields with <Name>{field_name}</Name> or <Name>{field_name}:XX</Name>")
            return mxl_content
        
        logger.info(f"Found {len(field_matches)} field(s) matching {field_name}")
        
        # Process each matching field
        for match_idx, field_match in enumerate(field_matches):
            # Get the full matched field
            field_content = field_match.group(0)
            original_field = field_content
            
            # Extract the actual field name for logging
            name_match = re.search(r'<Name>([^<]+)</Name>', field_content)
            actual_name = name_match.group(1) if name_match else field_name
            logger.info(f"Processing field #{match_idx + 1}: <Name>{actual_name}</Name>")
            
            # Remove <Link> tag completely (delete entire tag)
            before_link = '<Link>' in field_content
            field_content = re.sub(r'<Link>.*?</Link>\s*', '', field_content, flags=re.DOTALL)
            after_link = '<Link>' not in field_content
            logger.info(f"  Removed <Link> tag for {actual_name} - Before: {before_link}, After removed: {after_link}")
            
            # Remove <ImplicitRule> tag completely (delete entire tag)
            before_implicit = '<ImplicitRule>' in field_content
            field_content = re.sub(r'<ImplicitRule>.*?</ImplicitRule>\s*', '', field_content, flags=re.DOTALL)
            after_implicit = '<ImplicitRule>' not in field_content
            logger.info(f"  Removed <ImplicitRule> tag for {actual_name} - Before: {before_implicit}, After removed: {after_implicit}")
            
            # Comment out existing <ExplicitRule> content (only if not already commented)
            explicit_rule_match = re.search(r'<ExplicitRule>(.*?)</ExplicitRule>', field_content, re.DOTALL)
            if explicit_rule_match:
                existing_content = explicit_rule_match.group(1)
                if existing_content.strip():
                    # Comment each line (but skip lines that are already commented)
                    commented_lines = []
                    for line in existing_content.split('\n'):
                        if line.strip():
                            # Check if line is already commented (starts with //)
                            stripped_line = line.lstrip()
                            if not stripped_line.startswith('//'):
                                commented_lines.append('// ' + line)
                            else:
                                # Already commented, keep as is
                                commented_lines.append(line)
                        else:
                            commented_lines.append(line)
                    commented_content = '\n'.join(commented_lines)
                    field_content = field_content.replace(existing_content, commented_content)
                    logger.info(f"  Commented out existing ExplicitRule content for {actual_name}")
            
            # Build new ExplicitRule content (add # prefix to field_name in "into" clause)
            # Use column_d value directly (e.g., "SenderCode")
            new_rule = f'\nselect {text_value} into #{field_name} from codelist where name = "{column_c}" and {column_d} = {column_d};\n'
            
            # Add new rule to ExplicitRule
            if '<ExplicitRule>' in field_content and '</ExplicitRule>' in field_content:
                # Insert before closing tag
                field_content = field_content.replace('</ExplicitRule>', new_rule + '</ExplicitRule>')
            else:
                # Add ExplicitRule tags if they don't exist
                # Priority: After <Link> tag if exists, otherwise after </StoreLimit>
                if '<Link>' in field_content and '</Link>' in field_content:
                    # Add after </Link>, then remove the Link tag
                    field_content = re.sub(
                        r'(</Link>)',
                        r'\1\n<ExplicitRule>' + new_rule + '</ExplicitRule>',
                        field_content,
                        count=1
                    )
                    logger.info(f"  Added <ExplicitRule> after <Link> tag for {actual_name}")
                elif '</StoreLimit>' in field_content:
                    # Add after </StoreLimit>
                    field_content = re.sub(
                        r'(</StoreLimit>)',
                        r'\1\n<ExplicitRule>' + new_rule + '</ExplicitRule>',
                        field_content,
                        count=1
                    )
                    logger.info(f"  Added <ExplicitRule> after </StoreLimit> tag for {actual_name}")
                else:
                    # Fallback: Add before </Field>
                    field_content = field_content.replace('</Field>', f'<ExplicitRule>{new_rule}</ExplicitRule>\n</Field>')
                    logger.info(f"  Added <ExplicitRule> before </Field> tag for {actual_name}")
            
            logger.info(f"  Added new ExplicitRule for {actual_name}: {new_rule.strip()}")
            
            # Replace in original content using the regex pattern to ensure we replace the right field
            # Use the exact original field content for replacement to avoid replacing wrong fields
            mxl_content = mxl_content.replace(original_field, field_content, 1)
            logger.info(f"  Successfully replaced field content for {actual_name}")
        
        logger.info(f"Completed processing all {len(field_matches)} field(s) for {field_name}")
        
        return mxl_content
        
    except Exception as e:
        logger.error(f"Error processing SAP IDOC field: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return mxl_content

def process_sap_idoc_fields(mxl_file, field_configs):
    """
    Process all SAP IDOC field configurations
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read the MXL file
        with open(mxl_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Process each field configuration
        for config in field_configs:
            content = process_sap_idoc_field(content, config)
        
        # Write back to file
        # Debug: Check if SNDPRN field has the changes
        if '<Name>SNDPRN</Name>' in content:
            # Find the SNDPRN field in content
            sndprn_match = re.search(r'<Field>.*?<Name>SNDPRN</Name>.*?</Field>', content, re.DOTALL)
            if sndprn_match:
                sndprn_field = sndprn_match.group(0)
                if '<Link>' in sndprn_field:
                    logger.error("DEBUG: SNDPRN field still has Link tag in content!")
                else:
                    logger.info("DEBUG: SNDPRN field has no Link tag (correctly removed)")
                    
                if '<ImplicitRule>' in sndprn_field:
                    logger.error("DEBUG: SNDPRN field still has ImplicitRule tag in content!")
                else:
                    logger.info("DEBUG: SNDPRN field has no ImplicitRule tag (correctly removed)")
                    
                if '<ExplicitRule>' in sndprn_field and 'select Text6 into #SNDPRN' in sndprn_field:
                    logger.info("DEBUG: SNDPRN field has correct ExplicitRule in content")
                else:
                    logger.error("DEBUG: SNDPRN field does NOT have correct ExplicitRule in content!")
        
        with open(mxl_file, 'w', encoding='utf-8') as f:
            f.write(content)
            f.flush()  # Ensure write is flushed to disk
        
        logger.info(f"Successfully processed {len(field_configs)} SAP IDOC fields in {mxl_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error processing SAP IDOC fields: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def update_xml_output_settings(mxl_file, xml_config):
    """
    Update XML output settings in the <OUTPUT><XMLSyntax> section of MXL file
    
    Args:
        mxl_file: Path to MXL file
        xml_config: Dictionary with XML output configuration
    
    Returns:
        bool: True if successful, False otherwise
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read MXL file
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            content = f.read()
        
        # Find <OUTPUT><XMLSyntax> section
        xml_syntax_pattern = r'(<OUTPUT>.*?<XMLSyntax>)(.*?)(</XMLSyntax>.*?</OUTPUT>)'
        match = re.search(xml_syntax_pattern, content, re.DOTALL)
        
        if not match:
            logger.warning(f"Could not find <OUTPUT><XMLSyntax> section in {mxl_file}")
            return False
        
        opening = match.group(1)
        xml_syntax_content = match.group(2)
        closing = match.group(3)
        
        # Remove ONLY PublicID and SystemID that are directly before OutputHeader
        # This preserves PublicID/SystemID tags in other parts of XMLSyntax (like field definitions)
        # Pattern: Match PublicID/SystemID followed by optional whitespace, then OutputHeader
        # We need to match them together to ensure we only remove the ones before OutputHeader
        
        # Remove any existing PublicID/SystemID before OutputHeader (for idempotency)
        # Match: (optional PublicID)(optional SystemID)<OutputHeader>
        # Replace with: <OutputHeader>
        pattern = r'(?:<PublicID>.*?</PublicID>\s*)?(?:<SystemID>.*?</SystemID>\s*)?(?=<OutputHeader>)'
        xml_syntax_content = re.sub(pattern, '', xml_syntax_content, flags=re.DOTALL)
        
        # Now add them back if needed
        if xml_config['output_header'] == 'prolog_and_doctype':
            # Add PublicID and SystemID before OutputHeader
            tags_to_add = f'<PublicID>{xml_config["public_id"]}</PublicID>\n<SystemID>{xml_config["system_id"]}</SystemID>\n'
            
            if '<OutputHeader>' in xml_syntax_content:
                # Use regex to ensure we only replace the first occurrence
                xml_syntax_content = re.sub(r'<OutputHeader>', tags_to_add + '<OutputHeader>', xml_syntax_content, count=1)
            else:
                # Add at the beginning if OutputHeader doesn't exist
                xml_syntax_content = tags_to_add + xml_syntax_content
        # If not prolog_and_doctype, PublicID/SystemID stay removed (already done above)
        
        # Update OutputHeader
        if '<OutputHeader>' in xml_syntax_content:
            xml_syntax_content = re.sub(
                r'<OutputHeader>.*?</OutputHeader>',
                f'<OutputHeader>{xml_config["output_header"]}</OutputHeader>',
                xml_syntax_content
            )
        else:
            # Add OutputHeader if it doesn't exist
            xml_syntax_content = f'<OutputHeader>{xml_config["output_header"]}</OutputHeader>\n' + xml_syntax_content
        
        # Update OutputFormat
        if '<OutputFormat>' in xml_syntax_content:
            xml_syntax_content = re.sub(
                r'<OutputFormat>.*?</OutputFormat>',
                f'<OutputFormat>{xml_config["output_format"]}</OutputFormat>',
                xml_syntax_content
            )
        else:
            # Add OutputFormat if it doesn't exist
            xml_syntax_content = xml_syntax_content + f'\n<OutputFormat>{xml_config["output_format"]}</OutputFormat>'
        
        # Update UseCharEntityRef
        if '<UseCharEntityRef>' in xml_syntax_content:
            xml_syntax_content = re.sub(
                r'<UseCharEntityRef>.*?</UseCharEntityRef>',
                f'<UseCharEntityRef>{xml_config["use_char_entity_ref"]}</UseCharEntityRef>',
                xml_syntax_content
            )
        else:
            # Add UseCharEntityRef if it doesn't exist
            xml_syntax_content = xml_syntax_content + f'\n<UseCharEntityRef>{xml_config["use_char_entity_ref"]}</UseCharEntityRef>'
        
        # Reconstruct the content
        new_content = content.replace(match.group(0), opening + xml_syntax_content + closing)
        
        # Write back to file
        with open(mxl_file, 'w', encoding='UTF-8') as f:
            f.write(new_content)
        
        logger.info(f"Successfully updated XML output settings in {mxl_file}")
        logger.info(f"  OutputHeader: {xml_config['output_header']}")
        if xml_config['public_id']:
            logger.info(f"  PublicID: {xml_config['public_id']}")
        if xml_config['system_id']:
            logger.info(f"  SystemID: {xml_config['system_id']}")
        logger.info(f"  OutputFormat: {xml_config['output_format']}")
        logger.info(f"  UseCharEntityRef: {xml_config['use_char_entity_ref']}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating XML output settings: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False
def update_csv_output_settings(mxl_file, csv_config):
    """
    Update CSV output settings in the <OUTPUT><VarDelimSyntax> section of MXL file
    
    Args:
        mxl_file: Path to MXL file
        csv_config: Dictionary with CSV output configuration
    
    Returns:
        bool: True if successful, False otherwise
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read MXL file
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            content = f.read()
        
        # Find <OUTPUT><VarDelimSyntax> section
        vardelim_pattern = r'(<OUTPUT>.*?<VarDelimSyntax>)(.*?)(</VarDelimSyntax>.*?</OUTPUT>)'
        match = re.search(vardelim_pattern, content, re.DOTALL)
        
        if not match:
            logger.warning(f"Could not find <OUTPUT><VarDelimSyntax> section in {mxl_file}")
            return False
        
        opening = match.group(1)
        vardelim_content = match.group(2)
        closing = match.group(3)
        
        # Update FieldDelimiter
        if '<FieldDelimiter>' in vardelim_content:
            vardelim_content = re.sub(
                r'<FieldDelimiter>.*?</FieldDelimiter>',
                f'<FieldDelimiter>{csv_config["field_delimiter"]}</FieldDelimiter>',
                vardelim_content
            )
        else:
            # Add FieldDelimiter if it doesn't exist
            vardelim_content = f'<FieldDelimiter>{csv_config["field_delimiter"]}</FieldDelimiter>\n' + vardelim_content
        
        # Update QuoteCharacter
        if '<QuoteCharacter>' in vardelim_content:
            vardelim_content = re.sub(
                r'<QuoteCharacter>.*?</QuoteCharacter>',
                f'<QuoteCharacter>{csv_config["quote_character"]}</QuoteCharacter>',
                vardelim_content
            )
        else:
            # Add QuoteCharacter if it doesn't exist
            vardelim_content = vardelim_content + f'\n<QuoteCharacter>{csv_config["quote_character"]}</QuoteCharacter>'
        
        # Update IncludeColumnNames
        if '<IncludeColumnNames>' in vardelim_content:
            vardelim_content = re.sub(
                r'<IncludeColumnNames>.*?</IncludeColumnNames>',
                f'<IncludeColumnNames>{csv_config["include_column_names"]}</IncludeColumnNames>',
                vardelim_content
            )
        else:
            # Add IncludeColumnNames if it doesn't exist
            vardelim_content = vardelim_content + f'\n<IncludeColumnNames>{csv_config["include_column_names"]}</IncludeColumnNames>'
        
        # Reconstruct the content
        new_content = content.replace(match.group(0), opening + vardelim_content + closing)
        
        # Write back to file
        with open(mxl_file, 'w', encoding='UTF-8') as f:
            f.write(new_content)
        
        logger.info(f"Successfully updated CSV output settings in {mxl_file}")
        logger.info(f"  FieldDelimiter: {csv_config['field_delimiter']}")
        logger.info(f"  QuoteCharacter: {csv_config['quote_character']}")
        logger.info(f"  IncludeColumnNames: {csv_config['include_column_names']}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating CSV output settings: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def update_possyntax_delimiters(mxl_file, delimiter1, delimiter2):
    """
    Update PosSyntax section in MXL file with delimiter configuration
    
    Args:
        mxl_file: Path to MXL file
        delimiter1: First delimiter value (or empty string)
        delimiter2: Second delimiter value (or empty string)
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read the MXL file
        with open(mxl_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Find the <OUTPUT> section with <PosSyntax>
        output_pattern = r'(<OUTPUT>.*?<PosSyntax>)(.*?)(</PosSyntax>.*?</OUTPUT>)'
        output_match = re.search(output_pattern, content, re.DOTALL)
        
        if not output_match:
            logger.warning(f"Could not find <OUTPUT><PosSyntax> section in {mxl_file}")
            return False
        
        possyntax_content = output_match.group(2)
        logger.info(f"Found PosSyntax section")
        
        # Determine what to update based on delimiter values
        has_delimiter1 = bool(delimiter1)
        has_delimiter2 = bool(delimiter2)
        
        # Update DelimiterUsed
        if has_delimiter1:
            possyntax_content = re.sub(
                r'<DelimiterUsed>.*?</DelimiterUsed>',
                '<DelimiterUsed>Yes</DelimiterUsed>',
                possyntax_content
            )
            logger.info(f"Set DelimiterUsed to Yes")
            
            # Update Delimiter1 tag - convert self-closing to open/close and add value
            # Handle both <Delimiter1/> and <Delimiter1 /> (with or without space)
            if '<Delimiter1' in possyntax_content and '/>' in possyntax_content:
                possyntax_content = re.sub(
                    r'<Delimiter1\s*/>',
                    f'<Delimiter1>{delimiter1}</Delimiter1>',
                    possyntax_content
                )
            else:
                possyntax_content = re.sub(
                    r'<Delimiter1>.*?</Delimiter1>',
                    f'<Delimiter1>{delimiter1}</Delimiter1>',
                    possyntax_content
                )
            logger.info(f"Set Delimiter1 to {delimiter1}")
        else:
            possyntax_content = re.sub(
                r'<DelimiterUsed>.*?</DelimiterUsed>',
                '<DelimiterUsed>no</DelimiterUsed>',
                possyntax_content
            )
            logger.info(f"Set DelimiterUsed to no")
            
            # Clear Delimiter1 value if it exists
            if '<Delimiter1>' in possyntax_content and '</Delimiter1>' in possyntax_content:
                possyntax_content = re.sub(
                    r'<Delimiter1>.*?</Delimiter1>',
                    '<Delimiter1 />',
                    possyntax_content
                )
                logger.info(f"Cleared Delimiter1 value")
        
        # Update Delimiter2Used
        if has_delimiter2:
            possyntax_content = re.sub(
                r'<Delimiter2Used>.*?</Delimiter2Used>',
                '<Delimiter2Used>Yes</Delimiter2Used>',
                possyntax_content
            )
            logger.info(f"Set Delimiter2Used to Yes")
            
            # Update Delimiter2 tag - convert self-closing to open/close and add value
            # Handle both <Delimiter2/> and <Delimiter2 /> (with or without space)
            if '<Delimiter2' in possyntax_content and '/>' in possyntax_content:
                possyntax_content = re.sub(
                    r'<Delimiter2\s*/>',
                    f'<Delimiter2>{delimiter2}</Delimiter2>',
                    possyntax_content
                )
            else:
                possyntax_content = re.sub(
                    r'<Delimiter2>.*?</Delimiter2>',
                    f'<Delimiter2>{delimiter2}</Delimiter2>',
                    possyntax_content
                )
            logger.info(f"Set Delimiter2 to {delimiter2}")
        else:
            possyntax_content = re.sub(
                r'<Delimiter2Used>.*?</Delimiter2Used>',
                '<Delimiter2Used>no</Delimiter2Used>',
                possyntax_content
            )
            logger.info(f"Set Delimiter2Used to no")
            
            # Clear Delimiter2 value if it exists
            if '<Delimiter2>' in possyntax_content and '</Delimiter2>' in possyntax_content:
                possyntax_content = re.sub(
                    r'<Delimiter2>.*?</Delimiter2>',
                    '<Delimiter2 />',
                    possyntax_content
                )
                logger.info(f"Cleared Delimiter2 value")
        
        # Reconstruct the content
        new_content = content[:output_match.start()] + \
                     output_match.group(1) + \
                     possyntax_content + \
                     output_match.group(3) + \
                     content[output_match.end():]
        
        # Write back to file
        with open(mxl_file, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        logger.info(f"Successfully updated PosSyntax delimiters in {mxl_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error updating PosSyntax delimiters: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def add_syntax_record(mxl_file, syntax_type, offsets, remove_delimiters=False):
    """
    Add SyntaxRecord configuration to <INPUT><EDISyntax> section after <DecimalSeparator/>
    
    Args:
        mxl_file: Path to MXL file
        syntax_type: 'x12' or 'edifact'
        offsets: Dict with parsed offset values from Excel
        remove_delimiters: If True, set all delimiter "Used" tags to "no" and clear values
    
    Returns:
        bool: True if successful, False otherwise
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read MXL file
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            content = f.read()
        
        # Find <INPUT><EDISyntax> section
        input_edi_pattern = r'(<INPUT>.*?<EDISyntax>)(.*?)(</EDISyntax>.*?</INPUT>)'
        match = re.search(input_edi_pattern, content, re.DOTALL)
        
        if not match:
            logger.warning(f"Could not find <INPUT><EDISyntax> section in {mxl_file}")
            return False
        
        opening = match.group(1)
        edi_syntax_content = match.group(2)
        closing = match.group(3)
        
        # Handle delimiter removal if requested
        if remove_delimiters:
            logger.info("Removing delimiters (setting all to 'no' and clearing values)")
            # Set all delimiter "Used" tags to "no"
            edi_syntax_content = re.sub(r'<TagDelimiterUsed>yes</TagDelimiterUsed>', '<TagDelimiterUsed>no</TagDelimiterUsed>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<SegmentDelimiterUsed>yes</SegmentDelimiterUsed>', '<SegmentDelimiterUsed>no</SegmentDelimiterUsed>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<ElementDelimiterUsed>yes</ElementDelimiterUsed>', '<ElementDelimiterUsed>no</ElementDelimiterUsed>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<RepeatingElementDelimiterUsed>yes</RepeatingElementDelimiterUsed>', '<RepeatingElementDelimiterUsed>no</RepeatingElementDelimiterUsed>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<SubElementDelimiterUsed>yes</SubElementDelimiterUsed>', '<SubElementDelimiterUsed>no</SubElementDelimiterUsed>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<ReleaseCharacterUsed>yes</ReleaseCharacterUsed>', '<ReleaseCharacterUsed>no</ReleaseCharacterUsed>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<DecimalSeparatorUsed>yes</DecimalSeparatorUsed>', '<DecimalSeparatorUsed>no</DecimalSeparatorUsed>', edi_syntax_content)
            
            # Clear delimiter values (replace content between tags with empty)
            edi_syntax_content = re.sub(r'<TagDelimiter>[^<]*</TagDelimiter>', '<TagDelimiter></TagDelimiter>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<SegmentDelimiter>[^<]*</SegmentDelimiter>', '<SegmentDelimiter></SegmentDelimiter>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<ElementDelimiter>[^<]*</ElementDelimiter>', '<ElementDelimiter></ElementDelimiter>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<SubElementDelimiter>[^<]*</SubElementDelimiter>', '<SubElementDelimiter></SubElementDelimiter>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<ReleaseCharacter>[^<]*</ReleaseCharacter>', '<ReleaseCharacter></ReleaseCharacter>', edi_syntax_content)
            edi_syntax_content = re.sub(r'<DecimalSeparator>[^<]*</DecimalSeparator>', '<DecimalSeparator></DecimalSeparator>', edi_syntax_content)
        else:
            logger.info("Retaining existing delimiters")
        
        # Check if SyntaxRecord already exists
        if '<SyntaxRecordUsed>' in edi_syntax_content:
            logger.info("SyntaxRecord already exists, removing it first for clean insertion")
            # Remove existing SyntaxRecordUsed and SyntaxRecord tags
            edi_syntax_content = re.sub(
                r'<SyntaxRecordUsed>.*?</SyntaxRecordUsed>\s*',
                '',
                edi_syntax_content,
                flags=re.DOTALL
            )
            edi_syntax_content = re.sub(
                r'<SyntaxRecord>.*?</SyntaxRecord>\s*',
                '',
                edi_syntax_content,
                flags=re.DOTALL
            )
        
        # Build the SyntaxRecord XML based on type and offsets
        syntax_record_xml = f'''<SyntaxRecordUsed>yes</SyntaxRecordUsed>
<SyntaxRecord>
<Length>{offsets['length']}</Length>
<Tag>{offsets['tag']}</Tag>
<TagOffset>{offsets['tag_offset']}</TagOffset>
<ContainsData>{offsets['contains_data']}</ContainsData>
<TagDelimOffset>{offsets['tag_delim_offset']}</TagDelimOffset>
<SegmentDelimOffset>{offsets['segment_delim_offset']}</SegmentDelimOffset>
<ElementDelimOffset>{offsets['element_delim_offset']}</ElementDelimOffset>
<RepeatDelimOffset>{offsets['repeat_delim_offset']}</RepeatDelimOffset>
<SubElementDelimOffset>{offsets['subelement_delim_offset']}</SubElementDelimOffset>
<RelCharOffset>{offsets['rel_char_offset']}</RelCharOffset>
<DecimalSeparatorOffset>{offsets['decimal_sep_offset']}</DecimalSeparatorOffset>
</SyntaxRecord>'''
        
        # Find <DecimalSeparator> tag (with or without content) and insert after it
        # Handle both self-closing and content-based formats
        decimal_sep_pattern = r'<DecimalSeparator\s*/>'
        decimal_sep_content_pattern = r'<DecimalSeparator>[^<]*</DecimalSeparator>'
        
        if re.search(decimal_sep_pattern, edi_syntax_content):
            # Self-closing tag: <DecimalSeparator/>
            edi_syntax_content = re.sub(
                decimal_sep_pattern,
                lambda m: m.group(0) + '\n' + syntax_record_xml,
                edi_syntax_content,
                count=1
            )
            logger.info(f"Inserted {syntax_type.upper()} SyntaxRecord after <DecimalSeparator />")
        elif re.search(decimal_sep_content_pattern, edi_syntax_content):
            # Content-based tag: <DecimalSeparator>.</DecimalSeparator>
            edi_syntax_content = re.sub(
                decimal_sep_content_pattern,
                lambda m: m.group(0) + '\n' + syntax_record_xml,
                edi_syntax_content,
                count=1
            )
            logger.info(f"Inserted {syntax_type.upper()} SyntaxRecord after <DecimalSeparator>...</DecimalSeparator>")
        else:
            logger.warning("Could not find <DecimalSeparator> tag, adding SyntaxRecord at end of EDISyntax section")
            edi_syntax_content = edi_syntax_content + '\n' + syntax_record_xml
        
        # Reconstruct the content
        new_content = content.replace(match.group(0), opening + edi_syntax_content + closing)
        
        # Write back to file
        with open(mxl_file, 'w', encoding='UTF-8') as f:
            f.write(new_content)
        
        logger.info(f"Successfully added {syntax_type.upper()} SyntaxRecord to {mxl_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error adding syntax record: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def process_mxl_file(checklist_file, mxl_file):
    """
    Main processing function for a single MXL file
    
    Args:
        checklist_file: Path to Generic_checklistMain.xlsm
        mxl_file: Path to MXL file to process
    """
    logger = logging.getLogger(__name__)
    
    logger.info("=" * 80)
    logger.info(f"Processing MXL file: {mxl_file}")
    logger.info("=" * 80)
    
    # Check if this is an Inbound map by reading the MXL file
    try:
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            mxl_content = f.read()
        
        # Check for Inbound indicators in the MXL file
        # Look for <EDIAssociations_IN> with actual content (not empty tags)
        edi_in_match = re.search(r'<EDIAssociations_IN>(.*?)</EDIAssociations_IN>', mxl_content, re.DOTALL)
        if edi_in_match:
            edi_in_content = edi_in_match.group(1).strip()
            # Check if there's actual content (not just whitespace or empty tags)
            has_content = bool(re.search(r'<\w+>[^<\s]+</\w+>', edi_in_content))
            
            if not has_content:
                logger.info("This is an OUTBOUND map (EDIAssociations_IN is empty)")
                logger.info("Inbound Maps Features processing is only for INBOUND maps")
                logger.info("Skipping processing - no changes needed")
                return True  # Return success, as this is expected behavior
        else:
            logger.warning("Could not determine map direction, proceeding with caution")
    except Exception as e:
        logger.error(f"Error checking map direction: {e}")
        # Continue processing if we can't determine direction
    
    success = True
    
    # Process SAP IDOC fields first
    logger.info("Checking for SAP IDOC configuration...")
    sap_idoc_configs = read_sap_idoc_config(checklist_file)
    if sap_idoc_configs:
        logger.info(f"Found {len(sap_idoc_configs)} SAP IDOC field configurations")
        if not process_sap_idoc_fields(mxl_file, sap_idoc_configs):
            logger.error("Failed to process SAP IDOC fields")
            success = False
    else:
        logger.info("No SAP IDOC processing required")
    
    # Process XML Output Configuration
    logger.info("Checking for XML Output configuration...")
    xml_config = read_xml_output_config(checklist_file)
    if xml_config and xml_config['is_xml']:
        logger.info("XML Output configuration found and enabled (B18 = Yes)")
        if not update_xml_output_settings(mxl_file, xml_config):
            logger.error("Failed to update XML output settings")
            success = False
    else:
        logger.info("No XML Output processing required (B18 = No or not configured)")
    
    # Process CSV Output Configuration
    logger.info("Checking for CSV Output configuration...")
    csv_config = read_csv_output_config(checklist_file)
    if csv_config and csv_config['is_csv']:
        logger.info("CSV Output configuration found and enabled (B24 = Yes)")
        if not update_csv_output_settings(mxl_file, csv_config):
            logger.error("Failed to update CSV output settings")
            success = False
    else:
        logger.info("No CSV Output processing required (B24 = No or not configured)")
    
    # Process Syntax Record Configuration
    logger.info("Checking for Syntax Record configuration...")
    syntax_config = read_syntax_record_config(checklist_file)
    if syntax_config and syntax_config['syntax_type'] != 'no' and syntax_config.get('offsets'):
        logger.info(f"Syntax Record configuration found: {syntax_config['syntax_type'].upper()}")
        logger.info(f"Delimiter mode: {'Remove' if syntax_config.get('remove_delimiters') else 'Retain'}")
        
        # Determine the map type from the MXL file
        map_type = None
        try:
            # Check for X12 indicators
            if re.search(r'<AgencyID>X</AgencyID>', mxl_content, re.IGNORECASE):
                map_type = 'x12'
            # Check for EDIFACT indicators
            elif re.search(r'<AgencyID>E</AgencyID>', mxl_content, re.IGNORECASE):
                map_type = 'edifact'
            
            if map_type:
                logger.info(f"Detected map type from MXL file: {map_type.upper()}")
                
                # Only apply syntax record if map type matches Excel configuration
                if map_type == syntax_config['syntax_type']:
                    logger.info(f"Map type matches Excel configuration, applying {map_type.upper()} syntax record")
                    if not add_syntax_record(mxl_file, syntax_config['syntax_type'], syntax_config['offsets'], syntax_config.get('remove_delimiters', False)):
                        logger.error("Failed to add syntax record")
                        success = False
                else:
                    logger.info(f"Map type ({map_type.upper()}) does not match Excel configuration ({syntax_config['syntax_type'].upper()}), skipping syntax record")
            else:
                logger.warning("Could not determine map type from MXL file, skipping syntax record")
        except Exception as e:
            logger.error(f"Error determining map type: {e}")
            logger.info("Skipping syntax record due to error")
    else:
        logger.info("No Syntax Record processing required (B2 = No or not configured)")
    
    # Read positional configuration
    config = read_positional_config(checklist_file)
    if config is None:
        logger.error("Failed to read positional configuration")
        return False
    
    # Check if positional processing is enabled
    if not config['is_positional']:
        logger.info("Positional processing is disabled (B15 = No), skipping delimiter updates")
        return success
    
    logger.info("Positional processing is enabled (B15 = Yes)")
    
    # Check if any delimiters are provided
    has_delimiters = bool(config['delimiter1']) or bool(config['delimiter2'])
    
    if not has_delimiters:
        logger.info("No delimiters provided, setting DelimiterUsed and Delimiter2Used to 'no'")
        # Still need to update the file to set both to 'no'
        if not update_possyntax_delimiters(mxl_file, "", ""):
            return False
        return success
    
    # Process with provided delimiters
    logger.info(f"Processing with delimiters: Delimiter1='{config['delimiter1']}', Delimiter2='{config['delimiter2']}'")
    if not update_possyntax_delimiters(mxl_file, config['delimiter1'], config['delimiter2']):
        return False
    
    return success

def main():
    """Main entry point"""
    logger = setup_logging()
    
    if len(sys.argv) != 3:
        logger.error("Usage: python inbound_mapsFeatures.py <checklist_file> <mxl_file>")
        sys.exit(1)
    
    checklist_file = sys.argv[1]
    mxl_file = sys.argv[2]
    
    # Validate files exist
    if not os.path.exists(checklist_file):
        logger.error(f"Checklist file not found: {checklist_file}")
        sys.exit(1)
    
    if not os.path.exists(mxl_file):
        logger.error(f"MXL file not found: {mxl_file}")
        sys.exit(1)
    
    # Process the file
    success = process_mxl_file(checklist_file, mxl_file)
    
    if success:
        logger.info("Processing completed successfully")
        sys.exit(0)
    else:
        logger.error("Processing failed")
        sys.exit(1)

if __name__ == "__main__":
    main()

# Made with Bob
