import xml.etree.ElementTree as ET
import openpyxl
from pathlib import Path
import sys
import os
import re
import tempfile
import zipfile
import shutil

# Define namespace for the MXL files
NAMESPACE = {'ns': 'http://www.stercomm.com/SI/Map'}
DESCRIPTION_PATTERN = re.compile(r"(<Description>)(.*?)(</Description>)", re.IGNORECASE | re.DOTALL)
AUTHOR_PATTERN = re.compile(r"(<Author>)(.*?)(</Author>)", re.IGNORECASE | re.DOTALL)

def update_description_tag(text, new_value):
    """Update only the FIRST Description tag in MXL content"""
    return DESCRIPTION_PATTERN.sub(rf"\1{new_value}\3", text, count=1)

def update_author_tag(text, new_value="IBM"):
    """Update only the FIRST Author tag in MXL content"""
    return AUTHOR_PATTERN.sub(rf"\1{new_value}\3", text, count=1)

def extract_map_name(root):
    """Extract map name from MapDetails/Summary/Description"""
    description = root.find('.//ns:MapDetails/ns:Summary/ns:Description', NAMESPACE)
    if description is not None and description.text:
        return description.text.strip()
    
    map_details = root.find('.//ns:MapDetails', NAMESPACE)
    if map_details is not None:
        description = map_details.find('.//ns:Description', NAMESPACE)
        if description is not None and description.text:
            return description.text.strip()
    
    return "Unknown"

def get_format_from_tag(tag_name):
    """Convert tag name to format string"""
    format_map = {
        'XMLSyntax': 'XML',
        'EDISyntax': 'EDI',
        'PosSyntax': 'Positional',
        'VarDelimSyntax': 'CSV'
    }
    return format_map.get(tag_name, 'Unknown')

def check_sap_indicator(root, section_name=None):
    """
    Check if the map contains SAP indicator (EDI_DC40 in tag)
    
    Args:
        root: XML root element
        section_name: 'INPUT' or 'OUTPUT' to check specific section, None to check entire document
    
    Returns:
        bool: True if EDI_DC40 found in specified section or entire document
    """
    if section_name:
        # Check only in specified section (INPUT or OUTPUT)
        section = root.find(f'.//ns:{section_name}', NAMESPACE)
        if section is None:
            return False
        
        # Check in <Tag> elements within this section
        for tag_elem in section.findall('.//ns:Tag', NAMESPACE):
            if tag_elem.text and 'EDI_DC40' in tag_elem.text.strip():
                return True
        
        # Check in <XMLTag> elements within this section
        for xml_tag_elem in section.findall('.//ns:XMLTag', NAMESPACE):
            if xml_tag_elem.text and 'EDI_DC40' in xml_tag_elem.text.strip():
                return True
    else:
        # Check in entire document
        for tag_elem in root.findall('.//ns:Tag', NAMESPACE):
            if tag_elem.text and 'EDI_DC40' in tag_elem.text.strip():
                return True
        
        for xml_tag_elem in root.findall('.//ns:XMLTag', NAMESPACE):
            if xml_tag_elem.text and 'EDI_DC40' in xml_tag_elem.text.strip():
                return True
    
    return False

def extract_document_type_from_name(map_name):
    """
    Extract document type from map name after direction indicator.
    Example: SANOFI_BRA_ITECH_IB_SALESORDER -> SALESORDER
    """
    map_name_upper = map_name.upper()
    
    # Look for direction indicators and extract what comes after
    patterns = ['_IB_', '_OB_', '_I_', '_O_']
    
    for pattern in patterns:
        if pattern in map_name_upper:
            # Find the position after the pattern
            pos = map_name_upper.find(pattern) + len(pattern)
            # Get the rest of the string
            remainder = map_name[pos:]
            # Split by underscore and get first part (document type)
            if remainder:
                doc_type = remainder.split('_')[0]
                if doc_type:
                    return doc_type
    
    return None

def extract_edi_info(root, is_inbound, is_sap=False):
    """Extract EDI association information"""
    if is_inbound:
        edi_assoc = root.find('.//ns:EDIAssociations_IN', NAMESPACE)
    else:
        edi_assoc = root.find('.//ns:EDIAssociations_OUT', NAMESPACE)
    
    if edi_assoc is not None:
        binding_id = edi_assoc.find('.//ns:BindingID', NAMESPACE)
        agency_desc = edi_assoc.find('.//ns:AgencyDescription', NAMESPACE)
        version_id = edi_assoc.find('.//ns:VersionID', NAMESPACE)
        
        if binding_id is not None and binding_id.text:
            binding_text = binding_id.text.strip()
            agency_text = agency_desc.text.strip() if agency_desc is not None and agency_desc.text else ""
            version_text = version_id.text.strip() if version_id is not None and version_id.text else ""
            
            sap_suffix = ", SAP" if is_sap else ""
            
            if agency_text == "X12" or binding_text.isdigit():
                if version_text:
                    return f"EDI (X12, {binding_text}, {version_text}{sap_suffix})"
                else:
                    return f"EDI (X12, {binding_text}{sap_suffix})"
            else:
                if version_text:
                    return f"EDI (Edifact, {binding_text}, {version_text}{sap_suffix})"
                else:
                    return f"EDI (Edifact, {binding_text}{sap_suffix})"
    
    if is_inbound:
        section = root.find('.//ns:INPUT', NAMESPACE)
    else:
        section = root.find('.//ns:OUTPUT', NAMESPACE)
    
    if section is not None:
        group = section.find('.//ns:Group', NAMESPACE)
        if group is not None:
            name = group.find('ns:Name', NAMESPACE)
            if name is not None and name.text:
                name_text = name.text.strip()
                sap_suffix = ", SAP" if is_sap else ""
                if name_text.isdigit():
                    return f"EDI (X12, {name_text}{sap_suffix})"
                else:
                    return f"EDI (Edifact, {name_text}{sap_suffix})"
    
    sap_suffix = " (SAP)" if is_sap else ""
    return f"EDI{sap_suffix}"

def process_zip_file(zip_path, mxl_output_folder, zipfiles_folder="zipfiles"):
    """
    Extract MXL from ZIP (handles nested ZIPs), rename if needed, and save to output folder.
    New folder structure:
    - zipfiles/ : Stores nested ZIP files
    - old_mxlFiles/ : Immutable backup of extracted MXL files
    - mxl_files/ : Working directory for all updates
    
    Returns: (success, original_name, updated_name, mxl_file_path)
    """
    zip_name = os.path.splitext(os.path.basename(zip_path))[0]
    mxl_target_name = f"{zip_name}.mxl"
    original_description = None
    
    # Create zipfiles folder if it doesn't exist
    os.makedirs(zipfiles_folder, exist_ok=True)
    
    with tempfile.TemporaryDirectory() as temp_dir:
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # First, look for MXL files directly
        mxl_files = []
        nested_zips = []
        
        for root, _, files in os.walk(temp_dir):
            for file_name in files:
                file_path = os.path.join(root, file_name)
                if file_name.lower().endswith(".mxl"):
                    mxl_files.append(file_path)
                elif file_name.lower().endswith(".zip"):
                    nested_zips.append(file_path)
        
        # If no MXL files found directly, check nested ZIP files
        if not mxl_files and nested_zips:
            print(f"  📦 Found {len(nested_zips)} nested ZIP file(s), extracting...")
            
            # Save nested ZIPs to zipfiles folder
            for nested_zip in nested_zips:
                nested_zip_name = os.path.basename(nested_zip)
                dest_zip_path = os.path.join(zipfiles_folder, nested_zip_name)
                shutil.copy2(nested_zip, dest_zip_path)
            print(f"     💾 Saved {len(nested_zips)} ZIP file(s) to {zipfiles_folder}/")
            
            # Process nested ZIPs
            for nested_zip in nested_zips:
                nested_zip_name = os.path.splitext(os.path.basename(nested_zip))[0]
                with tempfile.TemporaryDirectory() as nested_temp_dir:
                    try:
                        with zipfile.ZipFile(nested_zip, "r") as nested_zip_ref:
                            nested_zip_ref.extractall(nested_temp_dir)
                        
                        # Look for MXL files in nested ZIP
                        for nested_root, _, nested_files in os.walk(nested_temp_dir):
                            for nested_file in nested_files:
                                if nested_file.lower().endswith(".mxl"):
                                    nested_mxl_path = os.path.join(nested_root, nested_file)
                                    
                                    # Process this MXL file
                                    with open(nested_mxl_path, "r", encoding="utf-8") as file:
                                        content = file.read()
                                    
                                    # Extract original description
                                    match = DESCRIPTION_PATTERN.search(content)
                                    nested_original_desc = None
                                    if match:
                                        nested_original_desc = match.group(2).strip()
                                    
                                    # Update author and description
                                    updated_content = update_author_tag(content, "IBM")
                                    updated_content = update_description_tag(updated_content, nested_zip_name)
                                    
                                    # Save to output folder (old_mxlFiles)
                                    os.makedirs(mxl_output_folder, exist_ok=True)
                                    output_mxl_path = os.path.join(mxl_output_folder, f"{nested_zip_name}.mxl")
                                    
                                    with open(output_mxl_path, "w", encoding="utf-8") as file:
                                        file.write(updated_content)
                                    
                                    print(f"     ✅ Extracted from nested ZIP: {nested_zip_name}.mxl")
                                    if nested_original_desc and nested_original_desc != nested_zip_name:
                                        print(f"        Base name: {nested_original_desc}")
                    except Exception as e:
                        print(f"     ⚠️  Error processing nested ZIP {nested_zip_name}: {e}")
            
            # After processing nested ZIPs, check if we got any MXL files
            mxl_files = []
            for root, _, files in os.walk(mxl_output_folder):
                for file_name in files:
                    if file_name.lower().endswith(".mxl"):
                        mxl_files.append(os.path.join(root, file_name))
            
            if mxl_files:
                # Return success for nested ZIP processing
                return True, None, zip_name, None
        
        # Process direct MXL files
        if not mxl_files:
            print(f"  ⚠️  Skipped {zip_path}: no .mxl file found")
            return False, None, None, None
        
        if len(mxl_files) == 1:
            # Single MXL file - process normally
            old_mxl_path = mxl_files[0]
            
            with open(old_mxl_path, "r", encoding="utf-8") as file:
                content = file.read()
            
            # Extract original description before any modification
            match = DESCRIPTION_PATTERN.search(content)
            if match:
                original_description = match.group(2).strip()
            
            # Update author tag to "IBM"
            updated_content = update_author_tag(content, "IBM")
            
            # Update description tag to match ZIP name
            updated_content = update_description_tag(updated_content, zip_name)
            
            # Save to output folder with ZIP name
            os.makedirs(mxl_output_folder, exist_ok=True)
            output_mxl_path = os.path.join(mxl_output_folder, mxl_target_name)
            
            with open(output_mxl_path, "w", encoding="utf-8") as file:
                file.write(updated_content)
            
            print(f"  ✅ Extracted and processed: {mxl_target_name}")
            if original_description and original_description != zip_name:
                print(f"     Base name: {original_description}")
            
            return True, original_description, zip_name, output_mxl_path
        else:
            # Multiple MXL files - process each one
            print(f"  📦 Found {len(mxl_files)} MXL files, processing all...")
            for mxl_file in mxl_files:
                mxl_file_name = os.path.splitext(os.path.basename(mxl_file))[0]
                
                with open(mxl_file, "r", encoding="utf-8") as file:
                    content = file.read()
                
                # Extract original description
                match = DESCRIPTION_PATTERN.search(content)
                file_original_desc = None
                if match:
                    file_original_desc = match.group(2).strip()
                
                # Update author and description
                updated_content = update_author_tag(content, "IBM")
                updated_content = update_description_tag(updated_content, mxl_file_name)
                
                # Save to output folder
                os.makedirs(mxl_output_folder, exist_ok=True)
                output_mxl_path = os.path.join(mxl_output_folder, f"{mxl_file_name}.mxl")
                
                with open(output_mxl_path, "w", encoding="utf-8") as file:
                    file.write(updated_content)
                
                print(f"     ✅ Processed: {mxl_file_name}.mxl")
                if file_original_desc and file_original_desc != mxl_file_name:
                    print(f"        Base name: {file_original_desc}")
            
            return True, None, zip_name, None

def parse_mxl_file(file_path):
    """Parse a single MXL file and extract mapping details"""
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        map_name = extract_map_name(root)
        
        input_section = root.find('.//ns:INPUT', NAMESPACE)
        output_section = root.find('.//ns:OUTPUT', NAMESPACE)
        
        if input_section is None or output_section is None:
            print(f"Warning: Could not find INPUT or OUTPUT section in {file_path}")
            return None
        
        input_format_tag = None
        for child in input_section:
            tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            if tag_name in ['XMLSyntax', 'EDISyntax', 'PosSyntax', 'VarDelimSyntax']:
                input_format_tag = tag_name
                break
        
        if not input_format_tag:
            if input_section.find('ns:XMLSyntax', NAMESPACE) is not None:
                input_format_tag = 'XMLSyntax'
            elif input_section.find('ns:EDISyntax', NAMESPACE) is not None:
                input_format_tag = 'EDISyntax'
            elif input_section.find('ns:PosSyntax', NAMESPACE) is not None:
                input_format_tag = 'PosSyntax'
            elif input_section.find('ns:VarDelimSyntax', NAMESPACE) is not None:
                input_format_tag = 'VarDelimSyntax'
        
        output_format_tag = None
        for child in output_section:
            tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            if tag_name in ['XMLSyntax', 'EDISyntax', 'PosSyntax', 'VarDelimSyntax']:
                output_format_tag = tag_name
                break
        
        if not output_format_tag:
            if output_section.find('ns:XMLSyntax', NAMESPACE) is not None:
                output_format_tag = 'XMLSyntax'
            elif output_section.find('ns:EDISyntax', NAMESPACE) is not None:
                output_format_tag = 'EDISyntax'
            elif output_section.find('ns:PosSyntax', NAMESPACE) is not None:
                output_format_tag = 'PosSyntax'
            elif output_section.find('ns:VarDelimSyntax', NAMESPACE) is not None:
                output_format_tag = 'VarDelimSyntax'
        
        if not input_format_tag or not output_format_tag:
            print(f"Warning: Could not determine input/output format in {file_path}")
            return None
        
        input_format = get_format_from_tag(input_format_tag)
        output_format = get_format_from_tag(output_format_tag)
        
        # Determine direction
        if input_format == 'EDI' or output_format == 'EDI':
            # If EDI is present, use EDI-based logic
            is_inbound = (input_format == 'EDI')
        else:
            # If no EDI on either side, check map name for direction indicators
            map_name_upper = map_name.upper()
            if '_I_' in map_name_upper or '_IB_' in map_name_upper or map_name_upper.startswith('IB_'):
                is_inbound = True
            elif '_O_' in map_name_upper or '_OB_' in map_name_upper or map_name_upper.startswith('OB_'):
                is_inbound = False
            else:
                # Default to inbound if no clear indicator
                is_inbound = True
        
        direction = "Inbound" if is_inbound else "Outbound"
        
        # Check for SAP indicator in INPUT and OUTPUT sections separately
        input_has_sap = False
        output_has_sap = False
        
        if input_format in ['Positional', 'XML']:
            input_has_sap = check_sap_indicator(root, 'INPUT')
        
        if output_format in ['Positional', 'XML']:
            output_has_sap = check_sap_indicator(root, 'OUTPUT')
        
        # Extract EDI information first
        if input_format == 'EDI':
            edi_info = extract_edi_info(root, True, False)
            input_format = edi_info
        elif output_format == 'EDI':
            edi_info = extract_edi_info(root, False, False)
            output_format = edi_info
        
        # For non-EDI maps, extract document type from map name
        doc_type = None
        if not input_format.startswith('EDI') and not output_format.startswith('EDI'):
            doc_type = extract_document_type_from_name(map_name)
        
        # Add SAP markers where EDI_DC40 is found
        if input_has_sap and input_format in ['Positional', 'XML', 'CSV']:
            input_format = f"{input_format} (SAP)"
        
        if output_has_sap and output_format in ['Positional', 'XML', 'CSV']:
            output_format = f"{output_format} (SAP)"
        
        # For non-EDI maps, add document type to the direction side (Inbound=INPUT, Outbound=OUTPUT)
        if doc_type:
            if is_inbound:
                # Inbound: Add document type to INPUT format (don't duplicate SAP)
                if '(SAP)' in input_format:
                    input_format = input_format.replace('(SAP)', f'({doc_type}, SAP)')
                else:
                    input_format = f"{input_format} ({doc_type})"
            else:
                # Outbound: Add document type to OUTPUT format (don't duplicate SAP)
                if '(SAP)' in output_format:
                    output_format = output_format.replace('(SAP)', f'({doc_type}, SAP)')
                else:
                    output_format = f"{output_format} ({doc_type})"
        
        return {
            'map_name': map_name,
            'direction': direction,
            'input_format': input_format,
            'output_format': output_format
        }
        
    except ET.ParseError as e:
        print(f"Error parsing {file_path}: {e}")
        return None
    except Exception as e:
        print(f"Unexpected error processing {file_path}: {e}")
        return None

def process_zip_folder(zip_folder_path, output_excel, mxl_output_folder):
    """Process all ZIP files, extract MXL files, and create Excel report"""
    print("=" * 70)
    print("MXL PARSER - Processing ZIP Files")
    print("=" * 70)
    
    zip_folder = Path(zip_folder_path)
    if not zip_folder.exists() or not zip_folder.is_dir():
        print(f"Error: Folder '{zip_folder_path}' does not exist")
        return
    
    zip_files = list(zip_folder.glob("*.zip"))
    if not zip_files:
        print(f"No .zip files found in {zip_folder_path}")
        return
    
    print(f"\nFound {len(zip_files)} ZIP file(s)")
    print("-" * 70)
    
    # Clean the output folder before processing
    mxl_output_path = Path(mxl_output_folder)
    if mxl_output_path.exists():
        print(f"\nCleaning output folder: {mxl_output_folder}")
        import shutil
        shutil.rmtree(mxl_output_path)
        print("✅ Old MXL files removed")
    
    os.makedirs(mxl_output_folder, exist_ok=True)
    
    processed_data = []
    for zip_file in sorted(zip_files):
        print(f"\nProcessing: {zip_file.name}")
        success, original_name, updated_name, mxl_path = process_zip_file(
            str(zip_file), mxl_output_folder
        )
        
        if success:
            if mxl_path:
                # Single MXL file case
                result = parse_mxl_file(mxl_path)
                if result:
                    processed_data.append({
                        'updated_name': updated_name,
                        'original_name': original_name if original_name else updated_name,
                        'direction': result['direction'],
                        'input_format': result['input_format'],
                        'output_format': result['output_format']
                    })
            else:
                # Multiple MXL files or nested ZIPs case - parse all files in output folder
                mxl_files = list(Path(mxl_output_folder).glob("*.mxl"))
                for mxl_file in mxl_files:
                    result = parse_mxl_file(mxl_file)
                    if result:
                        file_name = mxl_file.stem
                        processed_data.append({
                            'updated_name': file_name,
                            'original_name': result['map_name'],
                            'direction': result['direction'],
                            'input_format': result['input_format'],
                            'output_format': result['output_format']
                        })
    
    if not processed_data:
        print("\nNo data to write to Excel")
        return
    
    print("\n" + "=" * 70)
    print("Creating Excel Report")
    print("=" * 70)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping Details"
    
    ws['A1'] = "Map Name"
    ws['B1'] = "Direction"
    ws['C1'] = "Input Format"
    ws['D1'] = "Output Format"
    ws['E1'] = "Base Name"
    
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    row = 2
    for data in processed_data:
        ws[f'A{row}'] = data['updated_name']
        ws[f'B{row}'] = data['direction']
        ws[f'C{row}'] = data['input_format']
        ws[f'D{row}'] = data['output_format']
        ws[f'E{row}'] = data['original_name']
        row += 1
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_excel)
    print(f"\n✅ Excel saved: {output_excel}")
    print(f"✅ MXL files saved: {mxl_output_folder}")
    print(f"✅ Total maps: {len(processed_data)}")
    print("=" * 70)

def process_mxl_folder(mxl_folder_path, output_excel):
    """Process MXL files directly (no ZIP extraction) - includes Base Name column"""
    print("=" * 70)
    print("MXL PARSER - Processing MXL Files")
    print("=" * 70)
    
    mxl_folder = Path(mxl_folder_path)
    if not mxl_folder.exists() or not mxl_folder.is_dir():
        print(f"Error: Folder '{mxl_folder_path}' does not exist")
        return
    
    mxl_files = list(mxl_folder.glob("*.mxl"))
    if not mxl_files:
        print(f"No .mxl files found in {mxl_folder_path}")
        return
    
    print(f"\nFound {len(mxl_files)} MXL file(s)")
    print("-" * 70)
    
    processed_data = []
    for mxl_file in sorted(mxl_files):
        print(f"Processing: {mxl_file.name}")
        result = parse_mxl_file(mxl_file)
        
        if result:
            # For direct MXL files, Map Name and Base Name are the same
            map_name = result['map_name']
            processed_data.append({
                'map_name': map_name,
                'base_name': map_name,  # Same as map_name for direct MXL files
                'direction': result['direction'],
                'input_format': result['input_format'],
                'output_format': result['output_format']
            })
    
    if not processed_data:
        print("\nNo data to write to Excel")
        return
    
    print("\n" + "=" * 70)
    print("Creating Excel Report")
    print("=" * 70)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping Details"
    
    ws['A1'] = "Map Name"
    ws['B1'] = "Direction"
    ws['C1'] = "Input Format"
    ws['D1'] = "Output Format"
    ws['E1'] = "Base Name"
    
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    row = 2
    for data in processed_data:
        ws[f'A{row}'] = data['map_name']
        ws[f'B{row}'] = data['direction']
        ws[f'C{row}'] = data['input_format']
        ws[f'D{row}'] = data['output_format']
        ws[f'E{row}'] = data['base_name']
        row += 1
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_excel)
    print(f"\n✅ Excel saved: {output_excel}")
    print(f"✅ Total maps: {len(processed_data)}")
    print("=" * 70)

def main():
    """Main function with auto-detection of input type"""
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python mxl_parser.py <input_path> [output_excel] [mxl_output_folder]")
        print("\nSupported inputs:")
        print("  - Single ZIP file (with nested ZIPs or MXL files)")
        print("  - Folder containing ZIP files")
        print("  - Single MXL file")
        print("  - Folder containing MXL files")
        print("\nExamples:")
        print("  python mxl_parser.py /path/to/file.zip")
        print("  python mxl_parser.py /path/to/file.mxl")
        print("  python mxl_parser.py Renamezips/Allmaps")
        print("  python mxl_parser.py mxl_files")
        sys.exit(1)
    
    input_path = sys.argv[1]
    input_path_obj = Path(input_path)
    
    # Check if --mxl flag is present (for backward compatibility)
    is_mxl_mode = '--mxl' in sys.argv
    
    if not input_path_obj.exists():
        print(f"Error: Path '{input_path}' does not exist")
        sys.exit(1)
    
    output_excel = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] != '--mxl' else "mapping_results.xlsx"
    mxl_output_folder = sys.argv[3] if len(sys.argv) > 3 else "old_mxlFiles"
    
    # Auto-detect input type
    if input_path_obj.is_file():
        # Single file - check extension
        if input_path_obj.suffix.lower() == '.zip':
            # Single ZIP file - create temp folder and process
            print(f"Detected: Single ZIP file")
            temp_folder = Path(tempfile.mkdtemp())
            shutil.copy2(input_path, temp_folder / input_path_obj.name)
            process_zip_folder(str(temp_folder), output_excel, mxl_output_folder)
            shutil.rmtree(temp_folder)  # Clean up temp folder
            source_folder = Path(mxl_output_folder)
        elif input_path_obj.suffix.lower() == '.mxl':
            # Single MXL file - create temp folder and process
            print(f"Detected: Single MXL file")
            temp_folder = Path(tempfile.mkdtemp())
            shutil.copy2(input_path, temp_folder / input_path_obj.name)
            process_mxl_folder(str(temp_folder), output_excel)
            shutil.rmtree(temp_folder)  # Clean up temp folder
            source_folder = temp_folder
        else:
            print(f"Error: Unsupported file type '{input_path_obj.suffix}'")
            print("Supported types: .zip, .mxl")
            sys.exit(1)
    elif input_path_obj.is_dir():
        # Directory - check contents
        zip_files = list(input_path_obj.glob("*.zip"))
        mxl_files = list(input_path_obj.glob("*.mxl"))
        
        if is_mxl_mode or (mxl_files and not zip_files):
            # MXL folder mode
            print(f"Detected: Folder with MXL files")
            process_mxl_folder(input_path, output_excel)
            source_folder = Path(input_path)
        elif zip_files:
            # ZIP folder mode
            print(f"Detected: Folder with ZIP files")
            process_zip_folder(input_path, output_excel, mxl_output_folder)
            source_folder = Path(mxl_output_folder)
        else:
            print(f"Error: No ZIP or MXL files found in '{input_path}'")
            sys.exit(1)
    else:
        print(f"Error: '{input_path}' is not a valid file or directory")
        sys.exit(1)
    
    # Copy MXL files from old_mxlFiles (backup) to mxl_files (working directory)
    # old_mxlFiles = immutable backup, mxl_files = where all updates are applied
    dest_folder = Path('mxl_files')
    dest_folder.mkdir(exist_ok=True)
    
    # Clean mxl_files folder (remove old working files)
    print(f"\n📁 Preparing working directory: {dest_folder}")
    removed_count = 0
    for old_file in dest_folder.glob('*.mxl'):
        old_file.unlink()
        removed_count += 1
    if removed_count > 0:
        print(f"   ✅ Removed {removed_count} old file(s)")
    else:
        print("   ✅ No old files to remove")
    
    # Copy MXL files from backup to working directory
    print(f"\n📋 Copying MXL files from {source_folder} to {dest_folder}...")
    copied = 0
    
    for mxl_file in source_folder.glob('*.mxl'):
        dest_file = dest_folder / mxl_file.name
        shutil.copy2(mxl_file, dest_file)
        print(f"   ✅ Copied: {mxl_file.name}")
        copied += 1
    
    print(f"\n✅ Copied {copied} file(s) to working directory")
    print(f"💾 Original files backed up in: {source_folder}")
    print(f"🔧 Working files ready in: {dest_folder}")

if __name__ == "__main__":
    main()

# Made with Bob
