#!/usr/bin/env python3
"""
Process Data Manager - Extract and Apply Comment/Uncomment for Process Data Rules
Enhanced version with XML parsing and Standard Rules support

Usage:
  python process_data_manager.py extract mxl_files process_data_rules.xlsx
  python process_data_manager.py apply process_data_rules.xlsx mxl_files
"""

import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def extract_namespaces(root_element: ET.Element) -> Dict[str, str]:
    """Extract namespace mappings from XML root element."""
    namespaces = {}
    
    if root_element.tag.startswith('{'):
        namespace_uri = root_element.tag[1:root_element.tag.index('}')]
        
        for key, value in root_element.attrib.items():
            if value == namespace_uri:
                if key.startswith('{http://www.w3.org/2000/xmlns/}'):
                    prefix = key.split('}')[1]
                    namespaces[prefix] = value
                elif key == 'xmlns':
                    namespaces[''] = value
        
        if namespace_uri not in namespaces.values():
            if 'stercomm.com/mxl' in namespace_uri or 'sterling' in namespace_uri.lower():
                namespaces['mxl'] = namespace_uri
            else:
                namespaces[''] = namespace_uri
    
    for key, value in root_element.attrib.items():
        if key.startswith('{http://www.w3.org/2000/xmlns/}'):
            prefix = key.split('}')[1]
            if prefix not in namespaces:
                namespaces[prefix] = value
        elif key.startswith('xmlns:'):
            prefix = key.split(':', 1)[1]
            if prefix not in namespaces:
                namespaces[prefix] = value
        elif key == 'xmlns' and '' not in namespaces:
            namespaces[''] = value
    
    return namespaces


def extract_extended_rules(root: ET.Element, namespaces: Dict[str, str], map_name: str) -> List[Dict[str, Any]]:
    """Extract process data rules from Extended Rules (ExplicitRule elements)."""
    rules = []
    
    # Find all ExplicitRule elements
    explicit_rules = []
    for elem in root.iter():
        if elem.tag.endswith('ExplicitRule') or elem.tag == 'ExplicitRule':
            explicit_rules.append(elem)
    
    for rule_index, rule_element in enumerate(explicit_rules, start=1):
        # Check for child elements
        child_elements = list(rule_element)
        
        if child_elements:
            for child in child_elements:
                rule_text = child.text or ""
                if rule_text.strip():
                    rules.extend(_extract_from_rule_text(rule_text, map_name, "EXTENDED"))
        else:
            rule_text = rule_element.text or ""
            if rule_text.strip():
                rules.extend(_extract_from_rule_text(rule_text, map_name, "EXTENDED"))
    
    return rules


def _extract_from_rule_text(rule_text: str, map_name: str, rule_type: str) -> List[Dict[str, Any]]:
    """Extract process data references from rule text."""
    rules = []
    
    if not re.search(r'\bprocessdata\b', rule_text, re.IGNORECASE):
        return rules
    
    # Extract statements line-by-line
    statement_lines = []
    current_lines = []
    
    for raw_line in rule_text.splitlines():
        stripped_line = raw_line.lstrip()
        normalized_line = stripped_line[2:].lstrip() if stripped_line.startswith('//') else stripped_line
        
        if not current_lines and re.match(r'^(select|update)\b', normalized_line, re.IGNORECASE):
            current_lines = [raw_line]
        elif current_lines:
            current_lines.append(raw_line)
        
        if current_lines and ';' in raw_line:
            statement_lines.append('\n'.join(current_lines))
            current_lines = []
    
    xpath_pattern = re.compile(
        r'where\s+xpath\s*=\s*"([^"]+)"|where\s+xpath\s*=\s*\'([^\']+)\'',
        re.IGNORECASE | re.DOTALL
    )
    
    for statement_text in statement_lines:
        xpath_match = xpath_pattern.search(statement_text)
        if not xpath_match:
            continue
        
        xpath_value = xpath_match.group(1) if xpath_match.group(1) else xpath_match.group(2)
        xpath_result = _extract_xpathresult_value(statement_text)
        first_line = statement_text.splitlines()[0].lstrip()
        operation_line = first_line[2:].lstrip() if first_line.startswith('//') else first_line
        operation_match = re.match(r'^(select|update)\b', operation_line, re.IGNORECASE)
        
        if not operation_match:
            continue
        
        operation_type = operation_match.group(1).upper()
        comment_status = "//" if first_line.startswith('//') else ""
        
        rules.append({
            'Map Name': map_name,
            'Rule Type': rule_type,
            'Operation Type': 'UseSelect' if operation_type == 'SELECT' else 'UseUpdate',
            'Process Data XPath': xpath_value,
            'XPath Result': xpath_result,
            'Comment Status': comment_status
        })
    
    return rules


def _extract_xpathresult_value(statement_text: str) -> str:
    """Extract XPath Result value from statement."""
    # Check for SELECT ... INTO pattern
    into_match = re.search(
        r'select\s+xpathresult\s+into\s+(\w+)\s+from\s+processdata',
        statement_text,
        re.IGNORECASE | re.DOTALL
    )
    if into_match:
        return into_match.group(1).strip()
    
    # Fallback to UPDATE pattern
    match = re.search(
        r'xpathresult\s*=\s*(.+?)\s+where\s+xpath\b',
        statement_text,
        re.IGNORECASE | re.DOTALL
    )
    if not match:
        return ""
    
    return match.group(1).strip()


def extract_standard_rules(root: ET.Element, namespaces: Dict[str, str], map_name: str) -> List[Dict[str, Any]]:
    """Extract process data rules from Standard Rules (UseSelect/UseUpdate elements)."""
    rules = []
    
    # Build constant lookup tables
    constant_lookup = _build_constant_lookup(root)
    constant_index_lookup = _build_constant_index_lookup(root)
    
    # Find UseSelect elements
    useselect_elements = []
    for elem in root.iter():
        if elem.tag.endswith('UseSelect') or elem.tag == 'UseSelect':
            useselect_elements.append(elem)
    
    # Find UseUpdate elements
    useupdate_elements = []
    for elem in root.iter():
        if elem.tag.endswith('UseUpdate') or elem.tag == 'UseUpdate':
            useupdate_elements.append(elem)
    
    # Process UseSelect elements
    for use_element in useselect_elements:
        rule = _process_use_element(use_element, root, "UseSelect", map_name, namespaces, 
                                     constant_lookup, constant_index_lookup)
        if rule:
            rules.append(rule)
    
    # Process UseUpdate elements
    for use_element in useupdate_elements:
        rule = _process_use_element(use_element, root, "UseUpdate", map_name, namespaces,
                                     constant_lookup, constant_index_lookup)
        if rule:
            rules.append(rule)
    
    return rules


def _build_constant_lookup(root: ET.Element) -> Dict[int, str]:
    """Build lookup table of constant IDs to XPath values (numeric IDs only)."""
    constant_lookup = {}
    
    for elem in root.iter():
        tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag_name == 'Constant':
            value_elem = None
            constant_id = None
            
            for child in elem:
                child_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                if child_tag in ('Value', 'value'):
                    value_elem = child
                elif child_tag in ('ID', 'id', 'Id', 'ConstantID') and child.text:
                    try:
                        constant_id = int(child.text.strip())
                    except (TypeError, ValueError):
                        constant_id = None
            
            if constant_id is not None and value_elem is not None and value_elem.text:
                constant_lookup[constant_id] = value_elem.text.strip()
    
    return constant_lookup


def _build_string_constant_lookup(root: ET.Element) -> Dict[str, str]:
    """Build lookup table of string constant IDs to XPath values."""
    constant_lookup = {}
    
    for elem in root.iter():
        tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag_name == 'Constant':
            value_elem = None
            constant_id_str = None
            
            for child in elem:
                child_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                if child_tag in ('Value', 'value'):
                    value_elem = child
                elif child_tag in ('ID', 'id', 'Id', 'ConstantID') and child.text:
                    # Try to get as string (for non-numeric IDs like "Search Key 1")
                    constant_id_str = child.text.strip()
            
            if constant_id_str and value_elem is not None and value_elem.text:
                constant_lookup[constant_id_str] = value_elem.text.strip()
    
    return constant_lookup


def _build_constant_index_lookup(root: ET.Element) -> Dict[int, str]:
    """Build positional lookup table for constants."""
    constant_index_lookup = {}
    
    constants = []
    for elem in root.iter():
        tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag_name == 'Constant':
            constants.append(elem)
    
    for index, const_elem in enumerate(constants):
        for child in const_elem:
            child_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            if child_tag in ('Value', 'value') and child.text:
                constant_index_lookup[index] = child.text.strip()
                break
    
    return constant_index_lookup


def _process_use_element(use_element: ET.Element, root: ET.Element, element_type: str,
                         map_name: str, namespaces: Dict[str, str],
                         constant_lookup: Dict[int, str], 
                         constant_index_lookup: Dict[int, str]) -> Optional[Dict[str, Any]]:
    """Process a UseSelect or UseUpdate element."""
    # Check if this references processdata
    table_name = _get_child_text(use_element, 'TableName', namespaces)
    if not table_name:
        table_name = _get_child_text(use_element, 'tablename', namespaces)
    
    if not table_name or ('processdata' not in table_name.lower() and 'process data' not in table_name.lower()):
        return None
    
    # Extract SubTableNameConstantID
    subtable_constant_id = _get_child_text(use_element, 'SubTableNameConstantID', namespaces)
    if not subtable_constant_id:
        subtable_constant_id = _get_child_text(use_element, 'subtablenameconstantid', namespaces)
    
    if not subtable_constant_id:
        return None
    
    # Resolve XPath
    xpath_value = None
    const_id = None
    
    if subtable_constant_id.startswith('//'):
        try:
            const_id = int(subtable_constant_id[2:])
        except (ValueError, TypeError):
            const_id = None
    else:
        try:
            const_id = int(subtable_constant_id)
        except (ValueError, TypeError):
            const_id = None
    
    if const_id is not None:
        xpath_value = constant_index_lookup.get(const_id)
        if not xpath_value:
            xpath_value = constant_index_lookup.get(const_id + 1)
        if not xpath_value:
            xpath_value = constant_lookup.get(const_id)
    
    if not xpath_value:
        return None
    
    # Extract field name
    xpath_result = _extract_field_name(root, use_element, namespaces)
    
    # Determine comment status
    comment_status = _determine_comment_status(use_element)
    
    return {
        'Map Name': map_name,
        'Rule Type': 'STANDARD',
        'Operation Type': element_type,
        'Process Data XPath': xpath_value,
        'XPath Result': xpath_result,
        'Comment Status': comment_status
    }


def _get_child_text(element: ET.Element, tag_name: str, namespaces: Dict[str, str]) -> Optional[str]:
    """Get text content of a child element."""
    child = element.find(tag_name, namespaces)
    if child is not None and child.text:
        return child.text.strip()
    
    for prefix in namespaces.keys():
        if prefix:
            child = element.find(f'{prefix}:{tag_name}', namespaces)
            if child is not None and child.text:
                return child.text.strip()
    
    for child in element:
        child_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if child_tag == tag_name and child.text:
            return child.text.strip()
    
    return None


def _extract_field_name(root: ET.Element, use_element: ET.Element, namespaces: Dict[str, str]) -> str:
    """Extract field name from UseSelect/UseUpdate."""
    element_tag = use_element.tag.split('}')[-1] if '}' in use_element.tag else use_element.tag
    
    if element_tag == 'UseSelect':
        to_field_id = _extract_to_field_id(use_element, namespaces)
        if to_field_id:
            field_name = _find_field_name_by_id(root, to_field_id, namespaces)
            if field_name:
                return f"#{field_name}" if field_name.isdigit() else field_name
    
    field_element = _find_enclosing_field(root, use_element)
    if field_element is None:
        return ""
    
    name_value = _get_child_text(field_element, 'Name', namespaces)
    if not name_value:
        name_value = _get_child_text(field_element, 'name', namespaces)
    
    if not name_value:
        return ""
    
    return f"#{name_value}" if name_value.isdigit() else name_value


def _extract_to_field_id(use_element: ET.Element, namespaces: Dict[str, str]) -> Optional[str]:
    """Extract ToFieldID from UseSelect."""
    for mapping in use_element:
        mapping_tag = mapping.tag.split('}')[-1] if '}' in mapping.tag else mapping.tag
        if mapping_tag == 'Mapping':
            to_field_id = _get_child_text(mapping, 'ToFieldID', namespaces)
            if to_field_id:
                return to_field_id
    return None


def _find_field_name_by_id(root: ET.Element, field_id: str, namespaces: Dict[str, str]) -> Optional[str]:
    """Find Field element by ID and extract its Name."""
    for elem in root.iter():
        elem_tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if elem_tag == 'Field':
            id_value = _get_child_text(elem, 'ID', namespaces)
            if id_value == field_id:
                name_value = _get_child_text(elem, 'Name', namespaces)
                if name_value:
                    return name_value
    return None


def _find_enclosing_field(root: ET.Element, target_element: ET.Element) -> Optional[ET.Element]:
    """Find the Field element that encloses the target element."""
    for candidate in root.iter():
        tag_name = candidate.tag.split('}')[-1] if '}' in candidate.tag else candidate.tag
        if tag_name != 'Field':
            continue
        
        for descendant in candidate.iter():
            if descendant is target_element:
                return candidate
    
    return None


def _determine_comment_status(element: ET.Element) -> str:
    """Determine if element is commented."""
    for attr in ['disabled', 'commented', 'comment']:
        if attr in element.attrib:
            attr_value = element.attrib[attr].lower()
            if attr_value in ['true', 'yes', '1', 'commented']:
                return "//"
    return ""


def extract_process_data_rules(mxl_dir: str, output_excel: str) -> None:
    """Extract process data rules from MXL files to Excel using XML parsing."""
    print(f"\n{'='*80}")
    print("EXTRACTING PROCESS DATA RULES (Enhanced with XML Parsing)")
    print(f"{'='*80}\n")
    
    mxl_path = Path(mxl_dir)
    if not mxl_path.exists():
        print(f"Error: Directory not found: {mxl_dir}")
        sys.exit(1)
    
    mxl_files = list(mxl_path.glob('*.mxl'))
    if not mxl_files:
        print(f"No MXL files found in: {mxl_dir}")
        sys.exit(1)
    
    print(f"Found {len(mxl_files)} MXL files\n")
    
    all_rules = []
    
    for mxl_file in mxl_files:
        print(f"Processing: {mxl_file.name}")
        
        try:
            # Parse XML
            tree = ET.parse(str(mxl_file))
            root = tree.getroot()
            
            # Extract namespaces
            namespaces = extract_namespaces(root)
            
            # Extract from Extended Rules
            extended_rules = extract_extended_rules(root, namespaces, mxl_file.name)
            all_rules.extend(extended_rules)
            print(f"  - Found {len(extended_rules)} Extended Rule(s)")
            
            # Extract from Standard Rules
            standard_rules = extract_standard_rules(root, namespaces, mxl_file.name)
            all_rules.extend(standard_rules)
            print(f"  - Found {len(standard_rules)} Standard Rule(s)")
            
        except ET.ParseError as e:
            print(f"  ✗ XML parsing error: {e}")
            continue
        except Exception as e:
            print(f"  ✗ Error: {e}")
            continue
    
    if not all_rules:
        print("\nNo process data rules found in any MXL files.")
        sys.exit(0)
    
    # Create Excel file
    print(f"\nCreating Excel file: {output_excel}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Process_Data_Rules"
    
    # Headers
    headers = ['Map Name', 'Rule Type', 'Operation Type', 'Process Data XPath', 'XPath Result', 'Comment Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Data
    for row_idx, rule in enumerate(all_rules, 2):
        ws.cell(row=row_idx, column=1, value=rule['Map Name'])
        ws.cell(row=row_idx, column=2, value=rule['Rule Type'])
        ws.cell(row=row_idx, column=3, value=rule['Operation Type'])
        ws.cell(row=row_idx, column=4, value=rule['Process Data XPath'])
        ws.cell(row=row_idx, column=5, value=rule['XPath Result'])
        ws.cell(row=row_idx, column=6, value=rule['Comment Status'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    wb.save(output_excel)
    
    print(f"\n{'='*80}")
    print(f"EXTRACTION COMPLETE")
    print(f"Total rules extracted: {len(all_rules)}")
    print(f"  - Extended Rules: {sum(1 for r in all_rules if r['Rule Type'] == 'EXTENDED')}")
    print(f"  - Standard Rules: {sum(1 for r in all_rules if r['Rule Type'] == 'STANDARD')}")
    print(f"Excel file created: {output_excel}")
    print(f"{'='*80}\n")
    print("Next steps:")
    print("1. Open the Excel file")
    print("2. In 'Comment Status' column:")
    print("   - Put '//' to comment out (disable) the rule")
    print("   - Leave blank to uncomment (enable) the rule")
    print(f"3. Run: python {sys.argv[0]} apply {output_excel} {mxl_dir}")
    print()



def _apply_standard_rule_change(root: ET.Element, xpath: str, xpath_result: str, 
                                 desired_state: str, namespaces: Dict[str, str]) -> bool:
    """
    Apply comment/uncomment change to a Standard Rule.
    
    For Standard Rules, "comment" means removing the <ImplicitRuleDef> block.
    "Uncomment" is not supported as the deleted block cannot be reconstructed.
    
    Args:
        root: XML root element
        xpath: Process data XPath to match
        xpath_result: XPath result value to match
        desired_state: "//" for comment (remove), "" for uncomment (not supported)
        namespaces: Namespace context
    
    Returns:
        bool: True if change applied successfully, False otherwise
    """
    if desired_state != "//" and desired_state != "//":
        # Uncomment not supported for Standard Rules
        print(f"    ⚠ Warning: Uncomment not supported for Standard Rules (xpath: {xpath})")
        return False
    
    # Find all UseSelect and UseUpdate elements
    use_elements = []
    for elem in root.iter():
        tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag_name in ['UseSelect', 'UseUpdate']:
            use_elements.append(elem)
    
    # Find matching element
    for use_elem in use_elements:
        # Check if this references processdata
        table_name = _get_child_text(use_elem, 'TableName', namespaces)
        if not table_name or 'process' not in table_name.lower() or 'data' not in table_name.lower():
            continue
        
        # Get SubTableNameConstantID
        const_id_text = _get_child_text(use_elem, 'SubTableNameConstantID', namespaces)
        if not const_id_text:
            continue
        
        # Resolve the constant to get XPath
        const_id_text_stripped = const_id_text.strip()
        
        # Build constant lookups
        constant_lookup = _build_constant_lookup(root)
        constant_index_lookup = _build_constant_index_lookup(root)
        string_constant_lookup = _build_string_constant_lookup(root)
        
        # Try to resolve as numeric ID first
        resolved_xpath = None
        try:
            const_id = int(const_id_text_stripped)
            # Resolve XPath using numeric ID
            resolved_xpath = constant_index_lookup.get(const_id)
            if not resolved_xpath:
                resolved_xpath = constant_index_lookup.get(const_id + 1)
            if not resolved_xpath:
                resolved_xpath = constant_lookup.get(const_id)
        except (ValueError, TypeError):
            # Not a numeric ID, try string lookup
            resolved_xpath = string_constant_lookup.get(const_id_text_stripped)
        
        if not resolved_xpath:
            continue
        
        # Check if this matches our target
        if resolved_xpath == xpath:
            # Find and verify xpath_result matches
            field_elem = _find_enclosing_field(root, use_elem)
            if field_elem:
                field_name = _get_child_text(field_elem, 'Name', namespaces)
                if field_name:
                    # Normalize field name (add # if numeric)
                    normalized_name = f"#{field_name}" if field_name.isdigit() else field_name
                    if normalized_name == xpath_result or field_name == xpath_result:
                        # Found matching rule - remove ImplicitRuleDef
                        implicit_rule_def = _find_enclosing_implicit_rule_def(root, use_elem)
                        if implicit_rule_def:
                            return _remove_element_from_parent(root, implicit_rule_def)
    
    return False


def _find_enclosing_implicit_rule_def(root: ET.Element, target: ET.Element) -> Optional[ET.Element]:
    """Find the ImplicitRuleDef element that contains the target element."""
    for elem in root.iter():
        tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag_name == 'ImplicitRuleDef':
            # Check if target is a descendant
            for descendant in elem.iter():
                if descendant is target:
                    return elem
    return None


def _remove_element_from_parent(root: ET.Element, target: ET.Element) -> bool:
    """Remove an element from its parent in the XML tree."""
    for parent in root.iter():
        for child in list(parent):
            if child is target:
                parent.remove(child)
                return True
    return False


def _remove_namespaces(root: ET.Element) -> None:
    """Remove namespace prefixes from all elements in the XML tree."""
    for elem in root.iter():
        # Remove namespace from tag
        if '}' in elem.tag:
            elem.tag = elem.tag.split('}', 1)[1]
        # Remove namespace declarations from attributes
        attribs = elem.attrib
        for name, value in list(attribs.items()):
            if '}' in name:
                new_name = name.split('}', 1)[1]
                attribs[new_name] = value
                del attribs[name]


def apply_process_data_rules(input_excel: str, mxl_dir: str) -> None:
    """Apply comment/uncomment changes from Excel to MXL files."""
    print(f"\n{'='*80}")
    print("APPLYING PROCESS DATA RULE CHANGES (Enhanced with Standard Rules Support)")
    print(f"{'='*80}\n")
    
    # Read Excel file
    if not Path(input_excel).exists():
        print(f"Error: Excel file not found: {input_excel}")
        sys.exit(1)
    
    df = pd.read_excel(input_excel, sheet_name='Process_Data_Rules')
    print(f"Loaded {len(df)} rules from Excel\n")
    
    mxl_path = Path(mxl_dir)
    if not mxl_path.exists():
        print(f"Error: Directory not found: {mxl_dir}")
        sys.exit(1)
    
    # Group rules by map name and rule type
    changes_by_map = {}
    for _, row in df.iterrows():
        map_name = row['Map Name']
        rule_type = row.get('Rule Type', 'EXTENDED')  # Default to EXTENDED for backward compatibility
        xpath_value = row['Process Data XPath']
        xpath_result = row['XPath Result']
        new_status = str(row['Comment Status']).strip().lower()
        
        if map_name not in changes_by_map:
            changes_by_map[map_name] = []
        
        changes_by_map[map_name].append({
            'xpath': xpath_value,
            'result': xpath_result,
            'status': new_status,
            'rule_type': rule_type
        })
    
    # Process each MXL file
    total_changes = 0
    for map_name, changes in changes_by_map.items():
        mxl_file = mxl_path / map_name
        if not mxl_file.exists():
            print(f"Warning: MXL file not found: {map_name}")
            continue
        
        print(f"Processing: {map_name}")
        
        # Parse XML for this file
        try:
            tree = ET.parse(str(mxl_file))
            root = tree.getroot()
            namespaces = extract_namespaces(root)
        except ET.ParseError as e:
            print(f"  ✗ XML parsing error: {e}")
            continue
        
        # Read content for Extended Rules (text-based)
        with open(mxl_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        original_content = content
        xml_modified = False
        file_changes = 0
        
        for change in changes:
            xpath = change['xpath']
            result = change['result']
            new_status = change['status']
            rule_type = change.get('rule_type', 'EXTENDED')
            
            # Handle based on rule type
            if rule_type == 'STANDARD':
                # Handle Standard Rules (XML-based)
                success = _apply_standard_rule_change(root, xpath, result, new_status, namespaces)
                if success:
                    xml_modified = True
                    file_changes += 1
                    print(f"  ✓ {new_status or 'Uncommented'} Standard Rule: {xpath}")
                else:
                    print(f"  ✗ Failed to update Standard Rule: {xpath}")
            else:
                # Handle Extended Rules (text-based)
                # Escape special regex characters in xpath
                xpath_escaped = re.escape(xpath)
                # Handle result value - convert to string first to handle numeric values
                result_str = str(result)
                if result_str.startswith('"') and result_str.endswith('"'):
                    result_content = result_str[1:-1]
                    result_pattern = r'"' + re.escape(result_content) + r'"'
                else:
                    result_pattern = re.escape(result_str)
                
                if new_status == '//' or new_status == '//':
                    # Comment out - add // to each line
                    pattern = r'(?<!//\s)(?<!//)(\s*)(update\s+processdata\s+set\s+xpathresult\s*=\s*' + result_pattern + r'\s+where\s+xpath\s*=\s*"' + xpath_escaped + r'"\s*;)'
                    
                    # Debug: Check if pattern exists in content
                    if re.search(pattern, content, flags=re.IGNORECASE | re.DOTALL):
                        print(f"  [DEBUG] Found uncommented rule to comment: xpath={xpath}, result={result}")
                    else:
                        print(f"  [DEBUG] Rule already commented or not found: xpath={xpath}, result={result}")
                    
                    def comment_multiline(match):
                        statement = match.group(0)
                        commented_lines = []
                        for line in statement.split('\n'):
                            if line.strip():
                                leading_space = len(line) - len(line.lstrip())
                                commented_lines.append(line[:leading_space] + '//' + line[leading_space:])
                            else:
                                commented_lines.append(line)
                        return '\n'.join(commented_lines)
                    
                    new_content = re.sub(pattern, comment_multiline, content, flags=re.IGNORECASE | re.DOTALL)
                    
                    if new_content != content:
                        content = new_content
                        file_changes += 1
                        print(f"  ✓ Commented: {xpath}")
                    
                    # Also handle select statements
                    pattern = r'(?<!//\s)(?<!//)(\s*)(select\s+xpathresult\s+into\s+' + result_pattern + r'\s+from\s+processdata\s+where\s+xpath\s*=\s*"' + xpath_escaped + r'"\s*;)'
                    new_content = re.sub(pattern, comment_multiline, content, flags=re.IGNORECASE | re.DOTALL)
                    
                    if new_content != content:
                        content = new_content
                        file_changes += 1
                
                elif not new_status or new_status.strip() == '' or new_status == 'nan':
                    # Uncomment - remove // from each line
                    pattern = r'(update\s+processdata\s+set\s+xpathresult\s*=\s*' + result_pattern + r'\s+where\s+xpath\s*=\s*"' + xpath_escaped + r'"\s*;)'
                    
                    def uncomment_multiline(match):
                        statement = match.group(0)
                        lines = statement.split('\n')
                        uncommented_lines = []
                        for line in lines:
                            if line.strip():
                                uncommented_line = re.sub(r'^(\s*)//\s?', r'\1', line)
                                uncommented_lines.append(uncommented_line)
                            else:
                                uncommented_lines.append(line)
                        return '\n'.join(uncommented_lines)
                    
                    new_content = re.sub(pattern, uncomment_multiline, content, flags=re.IGNORECASE | re.DOTALL)
                    
                    if new_content != content:
                        content = new_content
                        file_changes += 1
                    
                    # Also handle select statements
                    pattern = r'(select\s+xpathresult\s+into\s+' + result_pattern + r'\s+from\s+processdata\s+where\s+xpath\s*=\s*"' + xpath_escaped + r'"\s*;)'
                    new_content = re.sub(pattern, uncomment_multiline, content, flags=re.IGNORECASE | re.DOTALL)
                    
                    if new_content != content:
                        content = new_content
                        file_changes += 1
        
        # Write back if changes were made
        if xml_modified:
            # Remove namespaces before writing
            _remove_namespaces(root)
            # Write XML changes
            tree.write(str(mxl_file), encoding='utf-8', xml_declaration=True)
            print(f"  ✓ Applied {file_changes} XML change(s)")
            total_changes += file_changes
        elif content != original_content:
            # Write text-based changes
            with open(mxl_file, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f"  ✓ Applied {file_changes} text change(s)")
            total_changes += file_changes
        else:
            print(f"  - No changes needed")
    
    print(f"\n{'='*80}")
    print(f"APPLICATION COMPLETE")
    print(f"Total changes applied: {total_changes}")
    print(f"{'='*80}\n")


def main():
    """Main function."""
    if len(sys.argv) < 3:
        print("Usage:")
        print(f"  Extract: python {sys.argv[0]} extract <mxl_directory> <output_excel>")
        print(f"  Apply:   python {sys.argv[0]} apply <input_excel> <mxl_directory>")
        print("\nExamples:")
        print(f"  python {sys.argv[0]} extract mxl_files process_data_rules.xlsx")
        print(f"  python {sys.argv[0]} apply process_data_rules.xlsx mxl_files")
        sys.exit(1)
    
    command = sys.argv[1].lower()
    
    if command == 'extract':
        if len(sys.argv) != 4:
            print("Error: Extract requires 2 arguments: <mxl_directory> <output_excel>")
            sys.exit(1)
        extract_process_data_rules(sys.argv[2], sys.argv[3])
    
    elif command == 'apply':
        if len(sys.argv) != 4:
            print("Error: Apply requires 2 arguments: <input_excel> <mxl_directory>")
            sys.exit(1)
        apply_process_data_rules(sys.argv[2], sys.argv[3])
    
    else:
        print(f"Error: Unknown command '{command}'. Use 'extract' or 'apply'.")
        sys.exit(1)


if __name__ == '__main__':
    main()

# Made with Bob
