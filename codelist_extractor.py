#!/usr/bin/env python3
"""
Codelist Extractor for MXL Files
Extracts codelist names from XML mapping files using two methods:
1. Extended Rule: Search for 'select...from codelist where NAME = "..."' patterns
2. Standard Rule: Search for trading partner codelists (if present in future)

Usage:
    python3 codelist_extractor.py file1.mxl file2.mxl file3.mxl
    python3 codelist_extractor.py *.mxl
    python3 codelist_extractor.py archive.zip
    python3 codelist_extractor.py  (processes all .mxl files in current directory)
"""

import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Set, List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys
import argparse
import zipfile
import tempfile
import shutil


class CodelistExtractor:
    def __init__(self):
        self.codelists_extended = set()
        self.codelists_standard = set()
        self.file_details = []
        
    def extract_extended_rule_codelists(self, file_path: str) -> Set[str]:
        """
        Extract codelists using Extended Rule:
        Search for patterns like: select ... from codelist where NAME = "CODELIST_NAME"
        """
        codelists = set()
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # Pattern to match: from codelist where NAME = "CODELIST_NAME"
            pattern = r'from\s+codelist\s+where\s+NAME\s*=\s*"([^"]+)"'
            matches = re.finditer(pattern, content, re.IGNORECASE)
            
            for match in matches:
                codelist_name = match.group(1)
                codelists.add(codelist_name)
                print(f"  Found codelist (Extended Rule): {codelist_name}")
                
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            
        return codelists
    
    def extract_standard_rule_codelists(self, file_path: str) -> Set[str]:
        """
        Extract codelists using Standard Rule:
        1. Find UseSelect with Trading Partner Code List
        2. Get SubTableNameConstantID value (e.g., 21)
        3. Find the (value + 1)th Constant in ConstantMap (e.g., 22nd)
        4. Extract codelist name from that Constant's Value
        """
        codelists = set()
        
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Find all UseSelect elements with Trading Partner references
            for elem in root.iter():
                if elem.tag.endswith('UseSelect'):
                    # Check if this UseSelect references Trading Partner
                    table_name = elem.find('.//{*}TableName')
                    if table_name is not None and table_name.text and 'Trading Partner' in table_name.text:
                        print(f"  Found UseSelect with Trading Partner: {table_name.text}")
                        
                        # Get SubTableNameConstantID
                        subtable_id_elem = elem.find('.//{*}SubTableNameConstantID')
                        if subtable_id_elem is not None and subtable_id_elem.text:
                            try:
                                subtable_id = int(subtable_id_elem.text.strip())
                                print(f"  SubTableNameConstantID: {subtable_id}")
                                
                                # Now find the ConstantMap and get the (subtable_id + 1)th Constant
                                constant_map = root.find('.//{*}ConstantMap')
                                if constant_map is not None:
                                    constants = constant_map.findall('.//{*}Constant')
                                    print(f"  Found {len(constants)} constants in ConstantMap")
                                    
                                    # Get the constant at position subtable_id + 1 (0-indexed, so just subtable_id)
                                    if subtable_id < len(constants):
                                        target_constant = constants[subtable_id]
                                        value_elem = target_constant.find('.//{*}Value')
                                        if value_elem is not None and value_elem.text:
                                            codelist_name = value_elem.text.strip()
                                            # Filter out non-codelist values
                                            if codelist_name and not codelist_name.startswith('http') and not codelist_name.startswith('Translation'):
                                                codelists.add(codelist_name)
                                                print(f"  Found codelist (Standard Rule): {codelist_name}")
                            except (ValueError, IndexError) as e:
                                print(f"  Error processing SubTableNameConstantID: {e}")
            
            # Alternative method: Look for UseUpdate with Process Data
            for elem in root.iter():
                if elem.tag.endswith('UseUpdate'):
                    table_name = elem.find('.//{*}TableName')
                    if table_name is not None and table_name.text and 'Process Data' in table_name.text:
                        print(f"  Found UseUpdate with Process Data")
                        
                        subtable_id_elem = elem.find('.//{*}SubTableNameConstantID')
                        if subtable_id_elem is not None and subtable_id_elem.text:
                            try:
                                subtable_id = int(subtable_id_elem.text.strip())
                                print(f"  SubTableNameConstantID: {subtable_id}")
                                
                                constant_map = root.find('.//{*}ConstantMap')
                                if constant_map is not None:
                                    constants = constant_map.findall('.//{*}Constant')
                                    
                                    if subtable_id < len(constants):
                                        target_constant = constants[subtable_id]
                                        value_elem = target_constant.find('.//{*}Value')
                                        if value_elem is not None and value_elem.text:
                                            codelist_name = value_elem.text.strip()
                                            if codelist_name and not codelist_name.startswith('http') and not codelist_name.startswith('Translation') and len(codelist_name) > 2:
                                                codelists.add(codelist_name)
                                                print(f"  Found codelist (Standard Rule - Process Data): {codelist_name}")
                            except (ValueError, IndexError) as e:
                                print(f"  Error processing: {e}")
            
            # Method 3: Search for constants with "Search Key" or codelist-like patterns
            constant_map = root.find('.//{*}ConstantMap')
            if constant_map is not None:
                constants = constant_map.findall('.//{*}Constant')
                for constant in constants:
                    const_id_elem = constant.find('.//{*}ConstantID')
                    value_elem = constant.find('.//{*}Value')
                    
                    if const_id_elem is not None and value_elem is not None:
                        const_id = const_id_elem.text
                        value = value_elem.text
                        
                        # Look for "Search Key" constants that might contain codelist names
                        if const_id and 'Search Key' in const_id and value:
                            value = value.strip()
                            # Filter for actual codelist names (not XPath expressions)
                            if not value.startswith('Translation') and not value.startswith('http') and '/' not in value and len(value) > 2:
                                codelists.add(value)
                                print(f"  Found codelist (Standard Rule - Search Key): {value}")
            
            
                
        except Exception as e:
            print(f"Error processing standard rule for {file_path}: {e}")
            
        return codelists
    
    def process_file(self, file_path: str) -> Dict:
        """Process a single MXL file and extract all codelists"""
        print(f"\nProcessing file: {file_path}")
        
        file_name = Path(file_path).name
        
        # Extract using Extended Rule
        extended_codelists = self.extract_extended_rule_codelists(file_path)
        
        # Extract using Standard Rule
        standard_codelists = self.extract_standard_rule_codelists(file_path)
        
        # Combine all codelists
        all_codelists = extended_codelists.union(standard_codelists)
        
        file_info = {
            'file_name': file_name,
            'extended_count': len(extended_codelists),
            'standard_count': len(standard_codelists),
            'total_count': len(all_codelists),
            'codelists': sorted(list(all_codelists))
        }
        
        self.codelists_extended.update(extended_codelists)
        self.codelists_standard.update(standard_codelists)
        self.file_details.append(file_info)
        
        return file_info
    
    def process_directory(self, directory: str = '.') -> None:
        """Process all MXL files in the directory"""
        path = Path(directory)
        mxl_files = list(path.glob('*.mxl'))
        
        if not mxl_files:
            print(f"No .mxl files found in {directory}")
            return
        
        print(f"Found {len(mxl_files)} MXL file(s)")
        
        for mxl_file in mxl_files:
            self.process_file(str(mxl_file))
    
    def process_zip_file(self, zip_path: str) -> None:
        """Extract and process all MXL files from a ZIP archive"""
        print(f"\nProcessing ZIP file: {zip_path}")
        
        if not zipfile.is_zipfile(zip_path):
            print(f"Error: {zip_path} is not a valid ZIP file")
            return
        
        # Create a temporary directory to extract files
        temp_dir = tempfile.mkdtemp(prefix='codelist_extract_')
        
        try:
            # Extract ZIP file
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
                print(f"Extracted ZIP contents to temporary directory")
            
            # Find all .mxl files in the extracted directory (including subdirectories)
            temp_path = Path(temp_dir)
            mxl_files = list(temp_path.rglob('*.mxl'))
            
            if not mxl_files:
                print(f"No .mxl files found in ZIP archive")
                return
            
            print(f"Found {len(mxl_files)} MXL file(s) in ZIP archive")
            
            # Process each MXL file
            for mxl_file in mxl_files:
                # Get relative path for better display
                rel_path = mxl_file.relative_to(temp_path)
                print(f"\nProcessing from ZIP: {rel_path}")
                self.process_file(str(mxl_file))
        
        finally:
            # Clean up temporary directory
            shutil.rmtree(temp_dir, ignore_errors=True)
            print(f"\nCleaned up temporary files")
    
    def generate_excel_report(self, output_file: str = 'codelist_report.xlsx') -> None:
        """Generate an Excel report with all extracted codelists"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create main codelist sheet with SI No, Map Name, Codelist Name
        ws_main = wb.create_sheet('Codelists')
        self._create_main_codelist_sheet(ws_main)
        
        # Save workbook
        wb.save(output_file)
        print(f"\n✓ Excel report generated: {output_file}")
    
    def _create_main_codelist_sheet(self, ws):
        """Create main sheet with SI No, Map Name, Codelist Name, and New Codelist Name"""
        # Header row with styling
        headers = ['SI No', 'Map Name', 'Codelist Name', 'New Codelist Name']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Collect all codelists with their source files
        codelist_entries = []
        for file_info in self.file_details:
            map_name = file_info['file_name']
            for codelist in file_info['codelists']:
                codelist_entries.append({
                    'map_name': map_name,
                    'codelist': codelist
                })
        
        # Sort by map name, then by codelist name
        codelist_entries.sort(key=lambda x: (x['map_name'], x['codelist']))
        
        # Add data rows
        for idx, entry in enumerate(codelist_entries, start=2):
            ws.cell(row=idx, column=1, value=idx-1)  # SI No
            ws.cell(row=idx, column=2, value=entry['map_name'])  # Map Name
            ws.cell(row=idx, column=3, value=entry['codelist'])  # Codelist Name
            ws.cell(row=idx, column=4, value='')  # New Codelist Name (empty for user input)
            
            # Add alternating row colors for better readability
            if idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=idx, column=col).fill = PatternFill(
                        start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
                    )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, ws):
        """Create summary sheet with overall statistics"""
        # Header
        ws['A1'] = 'Codelist Extraction Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Timestamp
        ws['A2'] = 'Generated:'
        ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Statistics
        row = 4
        ws[f'A{row}'] = 'Total Files Processed:'
        ws[f'B{row}'] = len(self.file_details)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total Unique Codelists (Extended Rule):'
        ws[f'B{row}'] = len(self.codelists_extended)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total Unique Codelists (Standard Rule):'
        ws[f'B{row}'] = len(self.codelists_standard)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total Unique Codelists (Combined):'
        all_codelists = self.codelists_extended.union(self.codelists_standard)
        ws[f'B{row}'] = len(all_codelists)
        ws[f'A{row}'].font = Font(bold=True)
        
        # File breakdown
        row += 3
        ws[f'A{row}'] = 'File Breakdown:'
        ws[f'A{row}'].font = Font(bold=True, underline='single')
        
        row += 1
        ws[f'A{row}'] = 'File Name'
        ws[f'B{row}'] = 'Codelists Found'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        for file_info in self.file_details:
            row += 1
            ws[f'A{row}'] = file_info['file_name']
            ws[f'B{row}'] = file_info['total_count']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _create_all_codelists_sheet(self, ws):
        """Create sheet with all unique codelists"""
        # Header
        ws['A1'] = 'All Unique Codelists'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        
        ws['A2'] = 'Codelist Name'
        ws['A2'].font = Font(bold=True)
        ws['A2'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Add all unique codelists
        all_codelists = sorted(list(self.codelists_extended.union(self.codelists_standard)))
        
        for idx, codelist in enumerate(all_codelists, start=3):
            ws[f'A{idx}'] = codelist
        
        ws.column_dimensions['A'].width = 50
    
    def _create_file_details_sheet(self, ws):
        """Create sheet with detailed file information"""
        # Header
        headers = ['File Name', 'Extended Rule Count', 'Standard Rule Count', 'Total Count', 'Codelists']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        
        # Add file details
        for row_idx, file_info in enumerate(self.file_details, start=2):
            ws.cell(row=row_idx, column=1, value=file_info['file_name'])
            ws.cell(row=row_idx, column=2, value=file_info['extended_count'])
            ws.cell(row=row_idx, column=3, value=file_info['standard_count'])
            ws.cell(row=row_idx, column=4, value=file_info['total_count'])
            ws.cell(row=row_idx, column=5, value=', '.join(file_info['codelists']))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 60
    
    def _create_extended_rule_sheet(self, ws):
        """Create sheet specifically for Extended Rule codelists"""
        # Header
        ws['A1'] = 'Extended Rule Codelists'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        
        ws['A2'] = 'Codelist Name'
        ws['B2'] = 'Pattern'
        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)
        ws['A2'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ws['B2'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Add extended rule codelists
        for idx, codelist in enumerate(sorted(list(self.codelists_extended)), start=3):
            ws[f'A{idx}'] = codelist
            ws[f'B{idx}'] = 'select...from codelist where NAME = "..."'
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50
    
    def print_summary(self):
        """Print a summary of extracted codelists"""
        print("\n" + "="*70)
        print("CODELIST EXTRACTION SUMMARY")
        print("="*70)
        
        print(f"\nTotal Files Processed: {len(self.file_details)}")
        print(f"Total Unique Codelists (Extended Rule): {len(self.codelists_extended)}")
        print(f"Total Unique Codelists (Standard Rule): {len(self.codelists_standard)}")
        
        all_codelists = self.codelists_extended.union(self.codelists_standard)
        print(f"Total Unique Codelists (Combined): {len(all_codelists)}")
        
        if self.codelists_extended:
            print("\n" + "-"*70)
            print("EXTENDED RULE CODELISTS:")
            print("-"*70)
            for codelist in sorted(self.codelists_extended):
                print(f"  • {codelist}")
        
        if self.codelists_standard:
            print("\n" + "-"*70)
            print("STANDARD RULE CODELISTS:")
            print("-"*70)
            for codelist in sorted(self.codelists_standard):
                print(f"  • {codelist}")
        
        print("\n" + "="*70)


def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(
        description='Extract codelist names from MXL files',
        epilog='''
Examples:
  python3 codelist_extractor.py file1.mxl file2.mxl
  python3 codelist_extractor.py *.mxl
  python3 codelist_extractor.py -o my_report.xlsx file1.mxl
  python3 codelist_extractor.py  (processes all .mxl files in current directory)
        ''',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        'files',
        nargs='*',
        help='MXL files or ZIP archives to process. If not specified, processes all .mxl files in current directory'
    )
    
    parser.add_argument(
        '-o', '--output',
        default='codelist_report.xlsx',
        help='Output Excel file name (default: codelist_report.xlsx)'
    )
    
    args = parser.parse_args()
    
    print("="*70)
    print("CODELIST EXTRACTOR FOR MXL FILES")
    print("="*70)
    
    extractor = CodelistExtractor()
    
    # Determine which files to process
    if args.files:
        # Process specified files
        print(f"\nProcessing {len(args.files)} specified file(s)")
        for file_path in args.files:
            if not Path(file_path).exists():
                print(f"Warning: File not found: {file_path}")
                continue
            
            # Check if it's a ZIP file
            if file_path.lower().endswith('.zip'):
                extractor.process_zip_file(file_path)
            else:
                extractor.process_file(file_path)
    else:
        # Process all MXL files in mxl_files directory
        print("\nNo files specified. Processing all .mxl files in mxl_files directory")
        extractor.process_directory('mxl_files')
    
    # Print summary
    extractor.print_summary()
    
    # Generate Excel report
    extractor.generate_excel_report(args.output)
    
    print("\n✓ Processing complete!")


if __name__ == '__main__':
    main()

# Made with Bob
