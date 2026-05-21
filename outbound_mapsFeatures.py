#!/usr/bin/env python3
"""
Outbound Maps Features Processor
Handles output delimiter configuration from Outbound_maps sheet in Generic_checklistMain.xlsm
"""

import sys
import os
import re
import logging
from datetime import datetime
import openpyxl

def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def read_output_delimiters_config(checklist_file):
    """
    Read output delimiters configuration from Outbound_maps sheet (rows 4-11)
    Returns: dict with delimiter configuration or None if "Keep as is" selected
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Outbound_maps']
        
        # Row 4, Column B: Output Delimiters dropdown (Keep as is / Defined)
        output_delimiters_cell = ws.cell(row=4, column=2).value
        output_delimiters_option = str(output_delimiters_cell).strip().lower() if output_delimiters_cell else "keep as is"
        
        # If "Keep as is", return None to skip processing
        if output_delimiters_option == "keep as is":
            wb.close()
            logger.info("Output Delimiters set to 'Keep as is' - skipping delimiter configuration")
            return None
        
        # If "Defined", read all delimiter values from rows 5-11
        # Based on Excel screenshot:
        # Row 5: Tag Delimiter
        # Row 6: Element Delimiter
        # Row 7: Repeating Element Delimiter
        # Row 8: Release Character
        # Row 9: Segment Delimiter
        # Row 10: Sub-Element Delimiter
        # Row 11: Decimal Separator
        
        def get_delimiter_value(row, col=2):
            """Get delimiter value, treating placeholders as blank"""
            cell = ws.cell(row=row, column=col)
            value = cell.value
            
            if value is None:
                return ""
            
            value_str = str(value).strip()
            
            # Check if it's ONLY the placeholder text (case insensitive)
            if value_str.lower() == "enter the delimiter":
                return ""
            
            # If it contains "Enter the delimiter" followed by a delimiter (e.g., "Enter the delimiter*")
            # Extract only the delimiter part
            if value_str.lower().startswith("enter the delimiter"):
                # Remove "Enter the delimiter" prefix and extract the actual delimiter
                delimiter_part = value_str[len("enter the delimiter"):].strip()
                # Remove common separators like ":", "-", etc. if present at start
                delimiter_part = delimiter_part.lstrip(':-').strip()
                return delimiter_part if delimiter_part else ""
            
            return value_str
        
        tag_delimiter = get_delimiter_value(5)
        element_delimiter = get_delimiter_value(6)
        repeating_element_delimiter = get_delimiter_value(7)
        release_character = get_delimiter_value(8)
        segment_delimiter = get_delimiter_value(9)
        sub_element_delimiter = get_delimiter_value(10)
        decimal_separator = get_delimiter_value(11)
        
        wb.close()
        
        config = {
            'is_defined': True,
            'tag_delimiter': tag_delimiter,
            'segment_delimiter': segment_delimiter,
            'element_delimiter': element_delimiter,
            'repeating_element_delimiter': repeating_element_delimiter,
            'sub_element_delimiter': sub_element_delimiter,
            'release_character': release_character,
            'decimal_separator': decimal_separator
        }
        
        logger.info(f"Output Delimiters config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading output delimiters config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
def read_positional_input_config(checklist_file):
    """
    Read positional input configuration from Outbound_maps sheet
    Returns: dict with 'is_positional', 'delimiter1', 'delimiter2'
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Outbound_maps']
        
        # Row 13, Column B: Is this a Positional? (Yes/No)
        is_positional_cell = ws.cell(row=13, column=2).value
        is_positional = str(is_positional_cell).strip().lower() if is_positional_cell else "no"
        
        # Row 13, Column C: Delimiter 1 (e.g., "cr")
        delimiter1_cell = ws.cell(row=13, column=3).value
        delimiter1 = ""
        if delimiter1_cell and str(delimiter1_cell).strip():
            cell_text = str(delimiter1_cell).strip()
            # Ignore placeholder text (grey italic in Excel)
            placeholder_patterns = ['delimiter 1', 'delimiter1', 'enter delimiter', 'placeholder']
            if not any(pattern in cell_text.lower() for pattern in placeholder_patterns):
                # Extract delimiter value after colon (e.g., "Delimiter 1: CR" -> "CR")
                if ':' in cell_text:
                    delimiter1 = cell_text.split(':', 1)[1].strip()
                else:
                    delimiter1 = cell_text
        
        # Row 14, Column C: Delimiter 2 (e.g., "lf")
        delimiter2_cell = ws.cell(row=14, column=3).value
        delimiter2 = ""
        if delimiter2_cell and str(delimiter2_cell).strip():
            cell_text = str(delimiter2_cell).strip()
            # Ignore placeholder text (grey italic in Excel)
            placeholder_patterns = ['delimiter 2', 'delimiter2', 'enter delimiter', 'placeholder']
            if not any(pattern in cell_text.lower() for pattern in placeholder_patterns):
                # Extract delimiter value after colon (e.g., "Delimiter 2: LF" -> "LF")
                if ':' in cell_text:
                    delimiter2 = cell_text.split(':', 1)[1].strip()
                else:
                    delimiter2 = cell_text
        
        wb.close()
        
        config = {
            'is_positional': is_positional == 'yes',
            'delimiter1': delimiter1,
            'delimiter2': delimiter2
        }
        
        logger.info(f"Positional input config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading positional input config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def read_csv_input_config(checklist_file):
    """
    Read CSV input configuration from Outbound_maps sheet
    Row 16: Is this a CSV? (Yes/No)
    Row 17: Input Delimiter (Default/Custom)
    Row 18: Field Delimiter
    Row 19: Quote Character
    Row 20: Include column names (Yes/No)
    Returns: dict with 'is_csv', 'field_delimiter', 'quote_character', 'include_column_names'
    """
    logger = logging.getLogger(__name__)
    
    try:
        wb = openpyxl.load_workbook(checklist_file, data_only=False)
        ws = wb['Outbound_maps']
        
        # Row 16, Column B: Is this a CSV? (Yes/No)
        is_csv_cell = ws.cell(row=16, column=2).value
        is_csv = str(is_csv_cell).strip().lower() if is_csv_cell else "no"
        
        # If not CSV, return early
        if is_csv != 'yes':
            wb.close()
            return {
                'is_csv': False,
                'field_delimiter': None,
                'quote_character': None,
                'include_column_names': None
            }
        
        # Row 17, Column B: Input Delimiter Type (Default/Custom)
        delimiter_type_cell = ws.cell(row=17, column=2).value
        delimiter_type = str(delimiter_type_cell).strip().lower() if delimiter_type_cell else "default"
        
        # Determine field delimiter and quote character
        if delimiter_type == "default":
            field_delimiter = ","
            quote_character = '"'
        else:  # Custom
            # Row 18, Column B: Field Delimiter
            field_delimiter_cell = ws.cell(row=18, column=2).value
            field_delimiter_str = str(field_delimiter_cell).strip() if field_delimiter_cell else ""
            # Skip placeholder text
            if field_delimiter_str.lower().startswith("enter"):
                field_delimiter = ","
            else:
                field_delimiter = field_delimiter_str if field_delimiter_str else ","
            
            # Row 19, Column B: Quote Character
            quote_character_cell = ws.cell(row=19, column=2).value
            quote_character_str = str(quote_character_cell).strip() if quote_character_cell else ""
            # Skip placeholder text
            if quote_character_str.lower().startswith("enter"):
                quote_character = '"'
            else:
                quote_character = quote_character_str if quote_character_str else '"'
        
        # Row 20, Column B: Include column names (Yes/No)
        include_names_cell = ws.cell(row=20, column=2).value
        include_names = str(include_names_cell).strip().lower() if include_names_cell else "no"
        include_column_names = "1" if include_names == "yes" else "0"
        
        wb.close()
        
        config = {
            'is_csv': True,
            'field_delimiter': field_delimiter,
            'quote_character': quote_character,
            'include_column_names': include_column_names
        }
        
        logger.info(f"CSV input config: {config}")
        return config
        
    except Exception as e:
        logger.error(f"Error reading CSV input config: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def update_input_possyntax_delimiters(mxl_file, delimiter1, delimiter2):
    """
    Update PosSyntax section in <INPUT> section of MXL file with delimiter configuration
    
    Args:
        mxl_file: Path to MXL file
        delimiter1: First delimiter value (or empty string)
        delimiter2: Second delimiter value (or empty string)
    """
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"=== Starting update_input_possyntax_delimiters ===")
        logger.info(f"Target file: {mxl_file}")
        logger.info(f"Delimiter1: '{delimiter1}', Delimiter2: '{delimiter2}'")
        
        # Read the MXL file
        with open(mxl_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        logger.info(f"File read successfully, content length: {len(content)} characters")
        
        # Find the <INPUT> section with <PosSyntax>
        input_pattern = r'(<INPUT>.*?<PosSyntax>)(.*?)(</PosSyntax>.*?</INPUT>)'
        input_match = re.search(input_pattern, content, re.DOTALL)
        
        if not input_match:
            logger.warning(f"Could not find <INPUT><PosSyntax> section in {mxl_file}")
            return False
        
        logger.info(f"Found INPUT/PosSyntax section at position {input_match.start()}-{input_match.end()}")
        
        possyntax_content = input_match.group(2)
        logger.info(f"Found PosSyntax section in INPUT")
        
        # Determine what to update based on delimiter values
        has_delimiter1 = bool(delimiter1)
        has_delimiter2 = bool(delimiter2)
        
        # Update DelimiterUsed
        if has_delimiter1:
            possyntax_content = re.sub(
                r'<DelimiterUsed>.*?</DelimiterUsed>',
                '<DelimiterUsed>Yes</DelimiterUsed>',
                possyntax_content
            )
            logger.info(f"Set DelimiterUsed to Yes")
            
            # Update Delimiter1 tag
            # Check if Delimiter1 is self-closing (not just if /> exists anywhere)
            if re.search(r'<Delimiter1\s*/>', possyntax_content):
                logger.info(f"Found self-closing Delimiter1 tag")
                old_content = possyntax_content
                possyntax_content = re.sub(
                    r'<Delimiter1\s*/>',
                    f'<Delimiter1>{delimiter1}</Delimiter1>',
                    possyntax_content
                )
                if old_content != possyntax_content:
                    logger.info(f"Replaced self-closing Delimiter1 tag")
            else:
                # Find what we're replacing
                match = re.search(r'<Delimiter1>(.*?)</Delimiter1>', possyntax_content)
                if match:
                    logger.info(f"Found Delimiter1 with value: '{match.group(1)}'")
                old_content = possyntax_content
                possyntax_content = re.sub(
                    r'<Delimiter1>.*?</Delimiter1>',
                    f'<Delimiter1>{delimiter1}</Delimiter1>',
                    possyntax_content
                )
                if old_content != possyntax_content:
                    logger.info(f"Successfully replaced Delimiter1 value")
                else:
                    logger.warning(f"Delimiter1 replacement had no effect!")
            logger.info(f"Set Delimiter1 to {delimiter1}")
        else:
            possyntax_content = re.sub(
                r'<DelimiterUsed>.*?</DelimiterUsed>',
                '<DelimiterUsed>no</DelimiterUsed>',
                possyntax_content
            )
            logger.info(f"Set DelimiterUsed to no")
            
            # Clear Delimiter1 value
            if re.search(r'<Delimiter1>.*?</Delimiter1>', possyntax_content):
                possyntax_content = re.sub(
                    r'<Delimiter1>.*?</Delimiter1>',
                    '<Delimiter1 />',
                    possyntax_content
                )
                logger.info(f"Cleared Delimiter1 value")
        
        # Update Delimiter2Used
        if has_delimiter2:
            possyntax_content = re.sub(
                r'<Delimiter2Used>.*?</Delimiter2Used>',
                '<Delimiter2Used>Yes</Delimiter2Used>',
                possyntax_content
            )
            logger.info(f"Set Delimiter2Used to Yes")
            
            # Update Delimiter2 tag
            # Check if Delimiter2 is self-closing (not just if /> exists anywhere)
            if re.search(r'<Delimiter2\s*/>', possyntax_content):
                logger.info(f"Found self-closing Delimiter2 tag")
                old_content = possyntax_content
                possyntax_content = re.sub(
                    r'<Delimiter2\s*/>',
                    f'<Delimiter2>{delimiter2}</Delimiter2>',
                    possyntax_content
                )
                if old_content != possyntax_content:
                    logger.info(f"Replaced self-closing Delimiter2 tag")
            else:
                # Find what we're replacing
                match = re.search(r'<Delimiter2>(.*?)</Delimiter2>', possyntax_content)
                if match:
                    logger.info(f"Found Delimiter2 with value: '{match.group(1)}'")
                old_content = possyntax_content
                possyntax_content = re.sub(
                    r'<Delimiter2>.*?</Delimiter2>',
                    f'<Delimiter2>{delimiter2}</Delimiter2>',
                    possyntax_content
                )
                if old_content != possyntax_content:
                    logger.info(f"Successfully replaced Delimiter2 value")
                else:
                    logger.warning(f"Delimiter2 replacement had no effect!")
            logger.info(f"Set Delimiter2 to {delimiter2}")
        else:
            possyntax_content = re.sub(
                r'<Delimiter2Used>.*?</Delimiter2Used>',
                '<Delimiter2Used>no</Delimiter2Used>',
                possyntax_content
            )
            logger.info(f"Set Delimiter2Used to no")
            
            # Clear Delimiter2 value
            if re.search(r'<Delimiter2>.*?</Delimiter2>', possyntax_content):
                possyntax_content = re.sub(
                    r'<Delimiter2>.*?</Delimiter2>',
                    '<Delimiter2 />',
                    possyntax_content
                )
                logger.info(f"Cleared Delimiter2 value")
        
        # Reconstruct the content
        new_content = content[:input_match.start()] + \
                     input_match.group(1) + \
                     possyntax_content + \
                     input_match.group(3) + \
                     content[input_match.end():]
        
        logger.info(f"Content reconstructed, new length: {len(new_content)} characters")
        logger.info(f"Writing changes back to file: {mxl_file}")
        
        # Write back to file
        with open(mxl_file, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        logger.info(f"File written successfully")
        logger.info(f"Successfully updated INPUT PosSyntax delimiters in {mxl_file}")
        logger.info(f"=== Completed update_input_possyntax_delimiters ===")
        return True
        
    except Exception as e:
        logger.error(f"Error updating INPUT PosSyntax delimiters: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def update_csv_input_settings(mxl_file, csv_config, clear_settings=False):
    """
    Update CSV input settings in the <INPUT><VarDelimSyntax> section of MXL file
    
    Args:
        mxl_file: Path to MXL file
        csv_config: Dictionary with CSV input configuration (can be None if clear_settings=True)
        clear_settings: If True, clear all CSV settings to defaults
    
    Returns:
        bool: True if successful, False otherwise
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read MXL file
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            content = f.read()
        
        # Find <INPUT><VarDelimSyntax> section
        vardelim_pattern = r'(<INPUT>.*?<VarDelimSyntax>)(.*?)(</VarDelimSyntax>.*?</INPUT>)'
        match = re.search(vardelim_pattern, content, re.DOTALL)
        
        if not match:
            logger.warning(f"Could not find <INPUT><VarDelimSyntax> section in {mxl_file}")
            return False
        
        # If clearing settings, use default values
        if clear_settings:
            csv_config = {
                'field_delimiter': ',',
                'quote_character': '"',
                'include_column_names': '0'
            }
            logger.info("Clearing CSV input settings to defaults")
        
        opening = match.group(1)
        vardelim_content = match.group(2)
        closing = match.group(3)
        
        # Update FieldDelimiter
        if '<FieldDelimiter>' in vardelim_content:
            vardelim_content = re.sub(
                r'<FieldDelimiter>.*?</FieldDelimiter>',
                f'<FieldDelimiter>{csv_config["field_delimiter"]}</FieldDelimiter>',
                vardelim_content
            )
        else:
            # Add FieldDelimiter if it doesn't exist
            vardelim_content = f'<FieldDelimiter>{csv_config["field_delimiter"]}</FieldDelimiter>\n' + vardelim_content
        
        # Update QuoteCharacter
        if '<QuoteCharacter>' in vardelim_content:
            vardelim_content = re.sub(
                r'<QuoteCharacter>.*?</QuoteCharacter>',
                f'<QuoteCharacter>{csv_config["quote_character"]}</QuoteCharacter>',
                vardelim_content
            )
        else:
            # Add QuoteCharacter if it doesn't exist
            vardelim_content = vardelim_content + f'\n<QuoteCharacter>{csv_config["quote_character"]}</QuoteCharacter>'
        
        # Update IncludeColumnNames
        if '<IncludeColumnNames>' in vardelim_content:
            vardelim_content = re.sub(
                r'<IncludeColumnNames>.*?</IncludeColumnNames>',
                f'<IncludeColumnNames>{csv_config["include_column_names"]}</IncludeColumnNames>',
                vardelim_content
            )
        else:
            # Add IncludeColumnNames if it doesn't exist
            vardelim_content = vardelim_content + f'\n<IncludeColumnNames>{csv_config["include_column_names"]}</IncludeColumnNames>'
        
        # Reconstruct the content
        new_content = content.replace(match.group(0), opening + vardelim_content + closing)
        
        # Write back to file
        with open(mxl_file, 'w', encoding='UTF-8') as f:
            f.write(new_content)
        
        logger.info(f"Successfully updated CSV input settings in {mxl_file}")
        logger.info(f"  FieldDelimiter: {csv_config['field_delimiter']}")
        logger.info(f"  QuoteCharacter: {csv_config['quote_character']}")
        logger.info(f"  IncludeColumnNames: {csv_config['include_column_names']}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating CSV input settings: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


def update_edi_output_delimiters(mxl_file, delimiter_config):
    """
    Update EDI output delimiters in the <OUTPUT><EDISyntax> section of MXL file
    
    Args:
        mxl_file: Path to MXL file
        delimiter_config: Dictionary with delimiter configuration
    
    Returns:
        bool: True if successful, False otherwise
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Read MXL file
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            content = f.read()
        
        # Find <OUTPUT><EDISyntax> section
        edi_syntax_pattern = r'(<OUTPUT>.*?<EDISyntax>)(.*?)(</EDISyntax>.*?</OUTPUT>)'
        match = re.search(edi_syntax_pattern, content, re.DOTALL)
        
        if not match:
            logger.warning(f"Could not find <OUTPUT><EDISyntax> section in {mxl_file}")
            return False
        
        opening = match.group(1)
        edi_syntax_content = match.group(2)
        closing = match.group(3)
        
        # Update each delimiter based on configuration
        # For each delimiter: if value exists, set Used=yes and update value; if blank, set Used=no
        
        # 1. Tag Delimiter
        if delimiter_config['tag_delimiter']:
            edi_syntax_content = re.sub(
                r'<TagDelimiterUsed>.*?</TagDelimiterUsed>',
                '<TagDelimiterUsed>yes</TagDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<TagDelimiter>.*?</TagDelimiter>',
                f'<TagDelimiter>{delimiter_config["tag_delimiter"]}</TagDelimiter>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<TagDelimiterUsed>.*?</TagDelimiterUsed>',
                '<TagDelimiterUsed>no</TagDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<TagDelimiter>.*?</TagDelimiter>',
                '<TagDelimiter></TagDelimiter>',
                edi_syntax_content
            )
        
        # 2. Segment Delimiter
        if delimiter_config['segment_delimiter']:
            edi_syntax_content = re.sub(
                r'<SegmentDelimiterUsed>.*?</SegmentDelimiterUsed>',
                '<SegmentDelimiterUsed>yes</SegmentDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<SegmentDelimiter>.*?</SegmentDelimiter>',
                f'<SegmentDelimiter>{delimiter_config["segment_delimiter"]}</SegmentDelimiter>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<SegmentDelimiterUsed>.*?</SegmentDelimiterUsed>',
                '<SegmentDelimiterUsed>no</SegmentDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<SegmentDelimiter>.*?</SegmentDelimiter>',
                '<SegmentDelimiter></SegmentDelimiter>',
                edi_syntax_content
            )
        
        # 3. Element Delimiter
        if delimiter_config['element_delimiter']:
            edi_syntax_content = re.sub(
                r'<ElementDelimiterUsed>.*?</ElementDelimiterUsed>',
                '<ElementDelimiterUsed>yes</ElementDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<ElementDelimiter>.*?</ElementDelimiter>',
                f'<ElementDelimiter>{delimiter_config["element_delimiter"]}</ElementDelimiter>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<ElementDelimiterUsed>.*?</ElementDelimiterUsed>',
                '<ElementDelimiterUsed>no</ElementDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<ElementDelimiter>.*?</ElementDelimiter>',
                '<ElementDelimiter></ElementDelimiter>',
                edi_syntax_content
            )
        
        # 4. Repeating Element Delimiter
        if delimiter_config['repeating_element_delimiter']:
            edi_syntax_content = re.sub(
                r'<RepeatingElementDelimiterUsed>.*?</RepeatingElementDelimiterUsed>',
                '<RepeatingElementDelimiterUsed>yes</RepeatingElementDelimiterUsed>',
                edi_syntax_content
            )
            # Handle both self-closing and regular closing tag formats
            edi_syntax_content = re.sub(
                r'<RepeatingElementDelimiter\s*/?>(?:.*?</RepeatingElementDelimiter>)?',
                f'<RepeatingElementDelimiter>{delimiter_config["repeating_element_delimiter"]}</RepeatingElementDelimiter>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<RepeatingElementDelimiterUsed>.*?</RepeatingElementDelimiterUsed>',
                '<RepeatingElementDelimiterUsed>no</RepeatingElementDelimiterUsed>',
                edi_syntax_content
            )
            # Handle both self-closing and regular closing tag formats
            edi_syntax_content = re.sub(
                r'<RepeatingElementDelimiter\s*/?>(?:.*?</RepeatingElementDelimiter>)?',
                '<RepeatingElementDelimiter />',
                edi_syntax_content
            )
        
        # 5. Sub-Element Delimiter
        if delimiter_config['sub_element_delimiter']:
            edi_syntax_content = re.sub(
                r'<SubElementDelimiterUsed>.*?</SubElementDelimiterUsed>',
                '<SubElementDelimiterUsed>yes</SubElementDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<SubElementDelimiter>.*?</SubElementDelimiter>',
                f'<SubElementDelimiter>{delimiter_config["sub_element_delimiter"]}</SubElementDelimiter>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<SubElementDelimiterUsed>.*?</SubElementDelimiterUsed>',
                '<SubElementDelimiterUsed>no</SubElementDelimiterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<SubElementDelimiter>.*?</SubElementDelimiter>',
                '<SubElementDelimiter></SubElementDelimiter>',
                edi_syntax_content
            )
        
        # 6. Release Character
        if delimiter_config['release_character']:
            edi_syntax_content = re.sub(
                r'<ReleaseCharacterUsed>.*?</ReleaseCharacterUsed>',
                '<ReleaseCharacterUsed>yes</ReleaseCharacterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<ReleaseCharacter>.*?</ReleaseCharacter>',
                f'<ReleaseCharacter>{delimiter_config["release_character"]}</ReleaseCharacter>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<ReleaseCharacterUsed>.*?</ReleaseCharacterUsed>',
                '<ReleaseCharacterUsed>no</ReleaseCharacterUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<ReleaseCharacter>.*?</ReleaseCharacter>',
                '<ReleaseCharacter></ReleaseCharacter>',
                edi_syntax_content
            )
        
        # 7. Decimal Separator
        if delimiter_config['decimal_separator']:
            edi_syntax_content = re.sub(
                r'<DecimalSeparatorUsed>.*?</DecimalSeparatorUsed>',
                '<DecimalSeparatorUsed>yes</DecimalSeparatorUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<DecimalSeparator>.*?</DecimalSeparator>',
                f'<DecimalSeparator>{delimiter_config["decimal_separator"]}</DecimalSeparator>',
                edi_syntax_content
            )
        else:
            edi_syntax_content = re.sub(
                r'<DecimalSeparatorUsed>.*?</DecimalSeparatorUsed>',
                '<DecimalSeparatorUsed>no</DecimalSeparatorUsed>',
                edi_syntax_content
            )
            edi_syntax_content = re.sub(
                r'<DecimalSeparator>.*?</DecimalSeparator>',
                '<DecimalSeparator></DecimalSeparator>',
                edi_syntax_content
            )
        
        # Reconstruct the content
        new_content = content.replace(match.group(0), opening + edi_syntax_content + closing)
        
        # Write back to file
        with open(mxl_file, 'w', encoding='UTF-8') as f:
            f.write(new_content)
        
        logger.info(f"Successfully updated EDI output delimiters in {mxl_file}")
        logger.info(f"  Tag Delimiter: {'yes' if delimiter_config['tag_delimiter'] else 'no'} - {delimiter_config['tag_delimiter']}")
        logger.info(f"  Segment Delimiter: {'yes' if delimiter_config['segment_delimiter'] else 'no'} - {delimiter_config['segment_delimiter']}")
        logger.info(f"  Element Delimiter: {'yes' if delimiter_config['element_delimiter'] else 'no'} - {delimiter_config['element_delimiter']}")
        logger.info(f"  Repeating Element Delimiter: {'yes' if delimiter_config['repeating_element_delimiter'] else 'no'} - {delimiter_config['repeating_element_delimiter']}")
        logger.info(f"  Sub-Element Delimiter: {'yes' if delimiter_config['sub_element_delimiter'] else 'no'} - {delimiter_config['sub_element_delimiter']}")
        logger.info(f"  Release Character: {'yes' if delimiter_config['release_character'] else 'no'} - {delimiter_config['release_character']}")
        logger.info(f"  Decimal Separator: {'yes' if delimiter_config['decimal_separator'] else 'no'} - {delimiter_config['decimal_separator']}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating EDI output delimiters: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def process_mxl_file(checklist_file, mxl_file):
    """
    Main processing function for a single MXL file
    
    Args:
        checklist_file: Path to Generic_checklistMain.xlsm
        mxl_file: Path to MXL file to process
    """
    logger = logging.getLogger(__name__)
    
    logger.info("=" * 80)
    logger.info(f"Processing MXL file: {mxl_file}")
    logger.info("=" * 80)
    
    # Check if this is an Outbound map with EDI output
    try:
        with open(mxl_file, 'r', encoding='UTF-8', errors='replace') as f:
            mxl_content = f.read()
        
        # Check for Outbound indicators
        edi_out_match = re.search(r'<EDIAssociations_OUT>(.*?)</EDIAssociations_OUT>', mxl_content, re.DOTALL)
        if edi_out_match:
            edi_out_content = edi_out_match.group(1).strip()
            has_content = bool(re.search(r'<\w+>[^<\s]+</\w+>', edi_out_content))
            
            if not has_content:
                logger.info("This is an INBOUND map (EDIAssociations_OUT is empty)")
                logger.info("Outbound Maps Features processing is only for OUTBOUND maps")
                logger.info("Skipping processing - no changes needed")
                return True
        else:
            logger.warning("Could not determine map direction, proceeding with caution")
        
        # Check if output is EDI format
        if '<OUTPUT>' not in mxl_content or '<EDISyntax>' not in mxl_content:
            logger.info("Output format is not EDI")
            logger.info("Output Delimiters configuration is only for EDI output")
            logger.info("Skipping processing - no changes needed")
            return True
            
    except Exception as e:
        logger.error(f"Error checking map direction and output format: {e}")
        return False
    
    success = True
    
    # Process Output Delimiters Configuration (only for EDI output)
    logger.info("Checking for Output Delimiters configuration...")
    delimiter_config = read_output_delimiters_config(checklist_file)
    if delimiter_config:
        logger.info("Output Delimiters configuration found and set to 'Defined'")
        if not update_edi_output_delimiters(mxl_file, delimiter_config):
            logger.error("Failed to update EDI output delimiters")
            success = False
    else:
        logger.info("No Output Delimiters processing required (Keep as is or not configured)")
    
    # Process Positional Input Configuration
    logger.info("Checking for Positional Input configuration...")
    positional_config = read_positional_input_config(checklist_file)
    if positional_config is None:
        logger.error("Failed to read positional input configuration")
        return False
    
    # Check if positional processing is enabled
    if positional_config['is_positional']:
        logger.info("Positional input processing is enabled (B15 = Yes)")
        
        # Check if any delimiters are provided
        has_delimiters = bool(positional_config['delimiter1']) or bool(positional_config['delimiter2'])
        
        if not has_delimiters:
            logger.info("No delimiters provided, setting DelimiterUsed and Delimiter2Used to 'no'")
            # Still need to update the file to set both to 'no'
            if not update_input_possyntax_delimiters(mxl_file, "", ""):
                logger.warning("Positional input update failed - file may not have PosSyntax section")
                # Don't fail the entire process if PosSyntax section is missing
        else:
            # Process with provided delimiters
            logger.info(f"Processing with delimiters: Delimiter1='{positional_config['delimiter1']}', Delimiter2='{positional_config['delimiter2']}'")
            if not update_input_possyntax_delimiters(mxl_file, positional_config['delimiter1'], positional_config['delimiter2']):
                logger.warning("Positional input update failed - file may not have PosSyntax section")
                # Don't fail the entire process if PosSyntax section is missing
    else:
        logger.info("Positional input processing is disabled (B13 = No)")
        logger.info("Clearing delimiters and setting DelimiterUsed and Delimiter2Used to 'no'")
        # Clear delimiters by passing empty strings
        if not update_input_possyntax_delimiters(mxl_file, "", ""):
            logger.warning("Positional input clear failed - file may not have PosSyntax section")
            # Don't fail the entire process if PosSyntax section is missing
    
    # Process CSV Input Configuration
    logger.info("Checking for CSV Input configuration...")
    csv_config = read_csv_input_config(checklist_file)
    if csv_config is None:
        logger.error("Failed to read CSV input configuration")
        return False
    
    if csv_config['is_csv']:
        logger.info("CSV Input configuration found and enabled (B16 = Yes)")
        if not update_csv_input_settings(mxl_file, csv_config):
            logger.warning("CSV input settings update failed - file may not have VarDelimSyntax section")
            # Don't fail the entire process if CSV section is missing
    else:
        logger.info("CSV Input processing is disabled (B16 = No)")
        logger.info("Clearing CSV settings to defaults (comma delimiter, double quote, no column names)")
        if not update_csv_input_settings(mxl_file, None, clear_settings=True):
            logger.warning("CSV input settings clear failed - file may not have VarDelimSyntax section")
            # Don't fail the entire process if CSV section is missing
    
    return success

def main():
    """Main entry point"""
    logger = setup_logging()
    
    if len(sys.argv) != 3:
        logger.error("Usage: python outbound_mapsFeatures.py <checklist_file> <mxl_file>")
        sys.exit(1)
    
    checklist_file = sys.argv[1]
    mxl_file = sys.argv[2]
    
    # Validate files exist
    if not os.path.exists(checklist_file):
        logger.error(f"Checklist file not found: {checklist_file}")
        sys.exit(1)
    
    if not os.path.exists(mxl_file):
        logger.error(f"MXL file not found: {mxl_file}")
        sys.exit(1)
    
    # Process the MXL file
    success = process_mxl_file(checklist_file, mxl_file)
    
    if success:
        logger.info("Processing completed successfully")
        sys.exit(0)
    else:
        logger.error("Processing failed")
        sys.exit(1)

if __name__ == "__main__":
    main()

# Made with Bob
